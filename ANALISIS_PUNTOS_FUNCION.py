"""
Script para generar análisis de Puntos de Función del proyecto ASEGURNASA
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear libro de Excel
wb = openpyxl.Workbook()

# ============================================================================
# HOJA 1: RESUMEN EJECUTIVO
# ============================================================================
ws_resumen = wb.active
ws_resumen.title = "Resumen Ejecutivo"

# Estilos
titulo_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
subtitulo_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
encabezado_fill = PatternFill(start_color="60a5fa", end_color="60a5fa", fill_type="solid")
total_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Título principal
ws_resumen['A1'] = 'ANÁLISIS DE PUNTOS DE FUNCIÓN'
ws_resumen['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_resumen['A1'].fill = titulo_fill
ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_resumen.merge_cells('A1:H1')
ws_resumen.row_dimensions[1].height = 30

# Información del proyecto
ws_resumen['A3'] = 'PROYECTO:'
ws_resumen['B3'] = 'ASEGURNASA - Sistema de Gestión de Pólizas de Seguros'
ws_resumen['A3'].font = Font(bold=True)
ws_resumen.merge_cells('B3:H3')

ws_resumen['A4'] = 'FECHA DE ANÁLISIS:'
ws_resumen['B4'] = '09/03/2026'
ws_resumen['A4'].font = Font(bold=True)

ws_resumen['A5'] = 'METODOLOGÍA:'
ws_resumen['B5'] = 'IFPUG (International Function Point Users Group) v4.3'
ws_resumen['A5'].font = Font(bold=True)
ws_resumen.merge_cells('B5:H5')

ws_resumen['A6'] = 'TIPO DE PROYECTO:'
ws_resumen['B6'] = 'Desarrollo Nuevo'
ws_resumen['A6'].font = Font(bold=True)

# Espacio
ws_resumen['A8'] = 'COMPONENTES DEL SISTEMA'
ws_resumen['A8'].font = Font(bold=True, size=12, color="FFFFFF")
ws_resumen['A8'].fill = subtitulo_fill
ws_resumen['A8'].alignment = Alignment(horizontal='center')
ws_resumen.merge_cells('A8:H8')

componentes_data = [
    ['Módulo', 'Descripción', 'LOC (Líneas de Código)'],
    ['main.py', 'Interface gráfica principal (customtkinter)', '~1,500'],
    ['database.py', 'Gestión de base de datos SQLite', '~300'],
    ['datos_mexico.py', 'Datos de estados y municipios', '~800'],
    ['TOTAL ESTIMADO', '', '~2,600']
]

row = 9
for i, data in enumerate(componentes_data):
    for col, value in enumerate(data, start=1):
        cell = ws_resumen.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
        elif i == len(componentes_data) - 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")

# Tecnologías
ws_resumen['A15'] = 'TECNOLOGÍAS UTILIZADAS'
ws_resumen['A15'].font = Font(bold=True, size=12, color="FFFFFF")
ws_resumen['A15'].fill = subtitulo_fill
ws_resumen['A15'].alignment = Alignment(horizontal='center')
ws_resumen.merge_cells('A15:H15')

tecnologias_data = [
    ['Categoría', 'Tecnología', 'Versión/Detalles'],
    ['Lenguaje', 'Python', '3.8+'],
    ['GUI Framework', 'customtkinter', '5.2.1'],
    ['Base de Datos', 'SQLite', '3.x'],
    ['Gestión de Excel', 'openpyxl', '3.x'],
    ['Gestión de PDF', 'reportlab', '4.x']
]

row = 16
for i, data in enumerate(tecnologias_data):
    for col, value in enumerate(data, start=1):
        cell = ws_resumen.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")

# Anchos de columna
ws_resumen.column_dimensions['A'].width = 25
ws_resumen.column_dimensions['B'].width = 50
ws_resumen.column_dimensions['C'].width = 20

# ============================================================================
# HOJA 2: ENTRADAS EXTERNAS (EI - External Inputs)
# ============================================================================
ws_ei = wb.create_sheet("Entradas Externas (EI)")

ws_ei['A1'] = 'ENTRADAS EXTERNAS (External Inputs - EI)'
ws_ei['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_ei['A1'].fill = titulo_fill
ws_ei['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_ei.merge_cells('A1:G1')
ws_ei.row_dimensions[1].height = 25

ws_ei['A3'] = 'Definición: Procesos que reciben datos desde fuera del sistema y actualizan archivos lógicos internos.'
ws_ei.merge_cells('A3:G3')
ws_ei['A3'].font = Font(italic=True)

# Tabla de EI
ei_headers = ['#', 'Nombre de la Función', 'DET', 'FTR', 'Complejidad', 'PF', 'Descripción']
row = 5
for col, header in enumerate(ei_headers, start=1):
    cell = ws_ei.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

ei_data = [
    [1, 'Agregar Nueva Póliza', 31, 1, 'Alta', 6, 'Formulario completo con 31 campos (datos personales, vehículo, póliza, coberturas, comentarios)'],
    [2, 'Actualizar Póliza Existente', 31, 1, 'Alta', 6, 'Modificación de póliza con todos los campos actualizables'],
    [3, 'Eliminar Póliza', 2, 1, 'Baja', 3, 'Eliminación con confirmación (ID + confirmación)'],
    [4, 'Agregar Fila Cobertura', 4, 1, 'Baja', 3, 'Fila dinámica en tabla (Cobertura, Suma, Deducible, Prima)'],
    [5, 'Eliminar Fila Cobertura', 1, 1, 'Baja', 3, 'Elimina fila específica de tabla de coberturas'],
    [6, 'Seleccionar Estado', 2, 1, 'Baja', 3, 'Selección de estado que activa cascada de municipios'],
    ['', 'TOTAL EI', '', '', '', 24, '']
]

row = 6
for i, data in enumerate(ei_data):
    for col, value in enumerate(data, start=1):
        cell = ws_ei.cell(row + i, col, value)
        cell.border = border
        if i == len(ei_data) - 1:  # Fila total
            cell.font = Font(bold=True)
            cell.fill = total_fill
            if col == 6:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in [3, 4, 6]:  # Columnas numéricas
            cell.alignment = Alignment(horizontal='center')

# Leyenda de complejidad
ws_ei['A14'] = 'COMPLEJIDAD Y VALORACIÓN:'
ws_ei['A14'].font = Font(bold=True)
ws_ei['A15'] = 'Baja = 3 PF | Media = 4 PF | Alta = 6 PF'
ws_ei.merge_cells('A15:G15')

ws_ei['A17'] = 'ABREVIATURAS:'
ws_ei['A17'].font = Font(bold=True)
ws_ei['A18'] = 'DET = Data Element Types (Elementos de Dato)'
ws_ei['A19'] = 'FTR = File Types Referenced (Archivos Referenciados)'
ws_ei['A20'] = 'PF = Puntos de Función'

# Anchos
ws_ei.column_dimensions['A'].width = 8
ws_ei.column_dimensions['B'].width = 30
ws_ei.column_dimensions['C'].width = 8
ws_ei.column_dimensions['D'].width = 8
ws_ei.column_dimensions['E'].width = 15
ws_ei.column_dimensions['F'].width = 8
ws_ei.column_dimensions['G'].width = 60

# ============================================================================
# HOJA 3: SALIDAS EXTERNAS (EO - External Outputs)
# ============================================================================
ws_eo = wb.create_sheet("Salidas Externas (EO)")

ws_eo['A1'] = 'SALIDAS EXTERNAS (External Outputs - EO)'
ws_eo['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_eo['A1'].fill = titulo_fill
ws_eo['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_eo.merge_cells('A1:G1')
ws_eo.row_dimensions[1].height = 25

ws_eo['A3'] = 'Definición: Procesos que envían datos fuera del sistema con lógica de procesamiento adicional.'
ws_eo.merge_cells('A3:G3')
ws_eo['A3'].font = Font(italic=True)

# Tabla de EO
eo_headers = ['#', 'Nombre de la Función', 'DET', 'FTR', 'Complejidad', 'PF', 'Descripción']
row = 5
for col, header in enumerate(eo_headers, start=1):
    cell = ws_eo.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

eo_data = [
    [1, 'Exportar a Excel', 31, 1, 'Alta', 7, 'Exporta pólizas seleccionadas con formato profesional, cálculos y múltiples hojas'],
    [2, 'Exportar a PDF', 31, 1, 'Alta', 7, 'Genera PDF con diseño profesional, logos, tablas formateadas'],
    [3, 'Calcular Días Vencimiento', 5, 1, 'Media', 5, 'Calcula días restantes, aplica lógica de color (rojo/amarillo/verde), valida fechas'],
    [4, 'Actualizar Lista Municipios', 3, 2, 'Media', 5, 'Filtra y carga municipios según estado seleccionado (cascada)'],
    [5, 'Generar Tarjeta Póliza', 8, 1, 'Baja', 4, 'Renderiza card en panel izquierdo con datos resumidos y formato condicional'],
    [6, 'Mostrar Alertas Vencimiento', 4, 1, 'Baja', 4, 'Muestra alertas visuales para pólizas próximas a vencer'],
    ['', 'TOTAL EO', '', '', '', 32, '']
]

row = 6
for i, data in enumerate(eo_data):
    for col, value in enumerate(data, start=1):
        cell = ws_eo.cell(row + i, col, value)
        cell.border = border
        if i == len(eo_data) - 1:
            cell.font = Font(bold=True)
            cell.fill = total_fill
            if col == 6:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in [3, 4, 6]:
            cell.alignment = Alignment(horizontal='center')

ws_eo['A15'] = 'COMPLEJIDAD Y VALORACIÓN:'
ws_eo['A15'].font = Font(bold=True)
ws_eo['A16'] = 'Baja = 4 PF | Media = 5 PF | Alta = 7 PF'
ws_eo.merge_cells('A16:G16')

# Anchos
ws_eo.column_dimensions['A'].width = 8
ws_eo.column_dimensions['B'].width = 30
ws_eo.column_dimensions['C'].width = 8
ws_eo.column_dimensions['D'].width = 8
ws_eo.column_dimensions['E'].width = 15
ws_eo.column_dimensions['F'].width = 8
ws_eo.column_dimensions['G'].width = 60

# ============================================================================
# HOJA 4: CONSULTAS EXTERNAS (EQ - External Inquiries)
# ============================================================================
ws_eq = wb.create_sheet("Consultas Externas (EQ)")

ws_eq['A1'] = 'CONSULTAS EXTERNAS (External Inquiries - EQ)'
ws_eq['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_eq['A1'].fill = titulo_fill
ws_eq['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_eq.merge_cells('A1:G1')
ws_eq.row_dimensions[1].height = 25

ws_eq['A3'] = 'Definición: Procesos que recuperan datos del sistema sin modificarlos ni aplicar lógica compleja.'
ws_eq.merge_cells('A3:G3')
ws_eq['A3'].font = Font(italic=True)

eq_headers = ['#', 'Nombre de la Función', 'DET', 'FTR', 'Complejidad', 'PF', 'Descripción']
row = 5
for col, header in enumerate(eq_headers, start=1):
    cell = ws_eq.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

eq_data = [
    [1, 'Buscar Pólizas (Tiempo Real)', 8, 1, 'Media', 4, 'Búsqueda multi-campo instantánea (nombre, RFC, póliza, placas, etc.)'],
    [2, 'Obtener Póliza por ID', 31, 1, 'Media', 4, 'Recupera todos los datos de una póliza específica'],
    [3, 'Listar Todas las Pólizas', 10, 1, 'Baja', 3, 'Muestra lista completa de pólizas en panel izquierdo'],
    [4, 'Cargar Coberturas de Póliza', 4, 1, 'Baja', 3, 'Deserializa JSON y carga tabla de coberturas'],
    [5, 'Verificar Vencimientos', 3, 1, 'Baja', 3, 'Consulta si hay pólizas próximas a vencer'],
    ['', 'TOTAL EQ', '', '', '', 17, '']
]

row = 6
for i, data in enumerate(eq_data):
    for col, value in enumerate(data, start=1):
        cell = ws_eq.cell(row + i, col, value)
        cell.border = border
        if i == len(eq_data) - 1:
            cell.font = Font(bold=True)
            cell.fill = total_fill
            if col == 6:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in [3, 4, 6]:
            cell.alignment = Alignment(horizontal='center')

ws_eq['A13'] = 'COMPLEJIDAD Y VALORACIÓN:'
ws_eq['A13'].font = Font(bold=True)
ws_eq['A14'] = 'Baja = 3 PF | Media = 4 PF | Alta = 6 PF'
ws_eq.merge_cells('A14:G14')

ws_eq.column_dimensions['A'].width = 8
ws_eq.column_dimensions['B'].width = 30
ws_eq.column_dimensions['C'].width = 8
ws_eq.column_dimensions['D'].width = 8
ws_eq.column_dimensions['E'].width = 15
ws_eq.column_dimensions['F'].width = 8
ws_eq.column_dimensions['G'].width = 60

# ============================================================================
# HOJA 5: ARCHIVOS LÓGICOS INTERNOS (ILF)
# ============================================================================
ws_ilf = wb.create_sheet("Archivos Internos (ILF)")

ws_ilf['A1'] = 'ARCHIVOS LÓGICOS INTERNOS (Internal Logical Files - ILF)'
ws_ilf['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_ilf['A1'].fill = titulo_fill
ws_ilf['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_ilf.merge_cells('A1:G1')
ws_ilf.row_dimensions[1].height = 25

ws_ilf['A3'] = 'Definición: Grupos de datos relacionados lógicamente, mantenidos dentro del sistema.'
ws_ilf.merge_cells('A3:G3')
ws_ilf['A3'].font = Font(italic=True)

ilf_headers = ['#', 'Nombre del Archivo', 'DET', 'RET', 'Complejidad', 'PF', 'Descripción']
row = 5
for col, header in enumerate(ilf_headers, start=1):
    cell = ws_ilf.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

ilf_data = [
    [1, 'Pólizas (polizas)', 31, 5, 'Alta', 15, 'Tabla principal: datos personales, póliza, vehículo, coberturas JSON, comentarios. Grupos: Personal(9), Póliza(10), Vehículo(10), Coberturas(1), Meta(2)'],
    ['', 'TOTAL ILF', '', '', '', 15, '']
]

row = 6
for i, data in enumerate(ilf_data):
    for col, value in enumerate(data, start=1):
        cell = ws_ilf.cell(row + i, col, value)
        cell.border = border
        if i == len(ilf_data) - 1:
            cell.font = Font(bold=True)
            cell.fill = total_fill
            if col == 6:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in [3, 4, 6]:
            cell.alignment = Alignment(horizontal='center')

ws_ilf['A9'] = 'COMPLEJIDAD Y VALORACIÓN:'
ws_ilf['A9'].font = Font(bold=True)
ws_ilf['A10'] = 'Baja = 7 PF | Media = 10 PF | Alta = 15 PF'
ws_ilf.merge_cells('A10:G10')

ws_ilf['A12'] = 'RET = Record Element Types (Subgrupos de datos dentro del archivo)'
ws_ilf['A12'].font = Font(italic=True)

ws_ilf.column_dimensions['A'].width = 8
ws_ilf.column_dimensions['B'].width = 25
ws_ilf.column_dimensions['C'].width = 8
ws_ilf.column_dimensions['D'].width = 8
ws_ilf.column_dimensions['E'].width = 15
ws_ilf.column_dimensions['F'].width = 8
ws_ilf.column_dimensions['G'].width = 70

# ============================================================================
# HOJA 6: ARCHIVOS DE INTERFAZ EXTERNA (EIF)
# ============================================================================
ws_eif = wb.create_sheet("Archivos Externos (EIF)")

ws_eif['A1'] = 'ARCHIVOS DE INTERFAZ EXTERNA (External Interface Files - EIF)'
ws_eif['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_eif['A1'].fill = titulo_fill
ws_eif['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_eif.merge_cells('A1:G1')
ws_eif.row_dimensions[1].height = 25

ws_eif['A3'] = 'Definición: Archivos de datos referenciados por el sistema pero mantenidos por otra aplicación.'
ws_eif.merge_cells('A3:G3')
ws_eif['A3'].font = Font(italic=True)

eif_headers = ['#', 'Nombre del Archivo', 'DET', 'RET', 'Complejidad', 'PF', 'Descripción']
row = 5
for col, header in enumerate(eif_headers, start=1):
    cell = ws_eif.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

eif_data = [
    [1, 'Catálogo Geográfico México', 35, 2, 'Baja', 5, 'datos_mexico.py: 32 estados + 500+ municipios. Grupos: Estados, Municipios por estado'],
    [2, 'Catálogos de Vehículos', 8, 3, 'Baja', 5, 'Listas predefinidas: Usos (8), Servicios (7), Tipos de moneda (3)'],
    ['', 'TOTAL EIF', '', '', '', 10, '']
]

row = 6
for i, data in enumerate(eif_data):
    for col, value in enumerate(data, start=1):
        cell = ws_eif.cell(row + i, col, value)
        cell.border = border
        if i == len(eif_data) - 1:
            cell.font = Font(bold=True)
            cell.fill = total_fill
            if col == 6:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in [3, 4, 6]:
            cell.alignment = Alignment(horizontal='center')

ws_eif['A10'] = 'COMPLEJIDAD Y VALORACIÓN:'
ws_eif['A10'].font = Font(bold=True)
ws_eif['A11'] = 'Baja = 5 PF | Media = 7 PF | Alta = 10 PF'
ws_eif.merge_cells('A11:G11')

ws_eif.column_dimensions['A'].width = 8
ws_eif.column_dimensions['B'].width = 30
ws_eif.column_dimensions['C'].width = 8
ws_eif.column_dimensions['D'].width = 8
ws_eif.column_dimensions['E'].width = 15
ws_eif.column_dimensions['F'].width = 8
ws_eif.column_dimensions['G'].width = 70

# ============================================================================
# HOJA 7: CÁLCULO TOTAL Y COSTEO
# ============================================================================
ws_total = wb.create_sheet("Cálculo Total y Costeo")

ws_total['A1'] = 'CÁLCULO TOTAL DE PUNTOS DE FUNCIÓN Y COSTEO'
ws_total['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_total['A1'].fill = titulo_fill
ws_total['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_total.merge_cells('A1:F1')
ws_total.row_dimensions[1].height = 25

# Tabla resumen PF
ws_total['A3'] = 'RESUMEN DE PUNTOS DE FUNCIÓN NO AJUSTADOS (UFP)'
ws_total['A3'].font = Font(bold=True, size=12)
ws_total.merge_cells('A3:F3')

resumen_headers = ['Componente', 'Cantidad', 'Puntos de Función', 'Subtotal PF']
row = 4
for col, header in enumerate(resumen_headers, start=1):
    cell = ws_total.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

resumen_data = [
    ['Entradas Externas (EI)', 6, '', 24],
    ['Salidas Externas (EO)', 6, '', 32],
    ['Consultas Externas (EQ)', 5, '', 17],
    ['Archivos Lógicos Internos (ILF)', 1, '', 15],
    ['Archivos de Interfaz Externa (EIF)', 2, '', 10],
    ['', '', 'TOTAL UFP', 98]
]

row = 5
for i, data in enumerate(resumen_data):
    for col, value in enumerate(data, start=1):
        cell = ws_total.cell(row + i, col, value)
        cell.border = border
        if i == len(resumen_data) - 1:
            cell.font = Font(bold=True, size=12)
            cell.fill = total_fill
            if col == 4:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
        if col in [2, 4]:
            cell.alignment = Alignment(horizontal='center')

# Factor de Ajuste de Valor (VAF)
ws_total['A13'] = 'FACTOR DE AJUSTE DE VALOR (VAF)'
ws_total['A13'].font = Font(bold=True, size=12)
ws_total.merge_cells('A13:F13')

ws_total['A14'] = 'Características Generales del Sistema:'
ws_total['A14'].font = Font(italic=True)
ws_total.merge_cells('A14:F14')

caracteristicas_headers = ['Característica', 'Nivel (0-5)', 'Justificación']
row = 15
for col, header in enumerate(caracteristicas_headers, start=1):
    cell = ws_total.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

caracteristicas_data = [
    ['1. Comunicación de datos', 2, 'Base de datos local, sin comunicación remota'],
    ['2. Procesamiento distribuido', 0, 'No aplica, aplicación standalone'],
    ['3. Rendimiento', 3, 'Búsqueda en tiempo real, cálculos instantáneos'],
    ['4. Configuración muy utilizada', 2, 'SQLite compartido, configuración estándar'],
    ['5. Tasa de transacciones', 3, 'CRUD completo con gestión de múltiples registros'],
    ['6. Entrada de datos en línea', 4, 'Formulario completo con 31 campos interactivos'],
    ['7. Eficiencia del usuario final', 5, 'UI profesional (customtkinter), búsqueda instantánea, validaciones'],
    ['8. Actualización en línea', 4, 'Actualización en tiempo real de listas y cálculos'],
    ['9. Complejidad de procesamiento', 4, 'Cálculos de fechas, cascadas, JSON, exportación Excel/PDF'],
    ['10. Reusabilidad', 2, 'Módulos separados (database.py, datos_mexico.py)'],
    ['11. Facilidad de instalación', 3, 'Scripts bat automáticos, requirements.txt'],
    ['12. Facilidad de operación', 5, 'Interfaz intuitiva, búsqueda automática, alertas visuales'],
    ['13. Múltiples sitios', 0, 'No diseñado para múltiples ubicaciones'],
    ['14. Facilidad de cambios', 4, 'Código modular, estructura clara, base de datos flexible'],
    ['TOTAL', 41, '']
]

row = 16
for i, data in enumerate(caracteristicas_data):
    for col, value in enumerate(data, start=1):
        cell = ws_total.cell(row + i, col, value)
        cell.border = border
        if i == len(caracteristicas_data) - 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
        if col == 2:
            cell.alignment = Alignment(horizontal='center')

ws_total.merge_cells('C16:C30')  # Fusionar celdas de justificación

# Cálculo VAF
ws_total['A32'] = 'Cálculo del VAF:'
ws_total['A32'].font = Font(bold=True)
ws_total['A33'] = 'VAF = 0.65 + (0.01 × Suma de Características)'
ws_total['A34'] = 'VAF = 0.65 + (0.01 × 41)'
ws_total['A35'] = 'VAF = 1.06'
ws_total['A35'].font = Font(bold=True, size=12)
ws_total['A35'].fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")

# Puntos de Función Ajustados
ws_total['A38'] = 'PUNTOS DE FUNCIÓN AJUSTADOS (AFP)'
ws_total['A38'].font = Font(bold=True, size=12, color="FFFFFF")
ws_total['A38'].fill = subtitulo_fill
ws_total.merge_cells('A38:F38')

ws_total['A39'] = 'AFP = UFP × VAF'
ws_total['A40'] = 'AFP = 98 × 1.06'
ws_total['A41'] = 'AFP = 103.88 ≈ 104 Puntos de Función'
ws_total['A41'].font = Font(bold=True, size=14)
ws_total['A41'].fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
ws_total.merge_cells('A41:F41')
ws_total['A41'].border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# COSTEO
ws_total['A44'] = 'ESTIMACIÓN DE ESFUERZO Y COSTO'
ws_total['A44'].font = Font(bold=True, size=12, color="FFFFFF")
ws_total['A44'].fill = subtitulo_fill
ws_total.merge_cells('A44:F44')

ws_total['A46'] = 'Parámetros de Conversión:'
ws_total['A46'].font = Font(bold=True)

ws_total['A47'] = 'Horas por PF (industria):'
ws_total['B47'] = '6-8 horas/PF'
ws_total['A48'] = 'Horas estimadas (usando 7 h/PF):'
ws_total['B48'] = '104 × 7 = 728 horas'
ws_total['B48'].font = Font(bold=True)

ws_total['A50'] = 'Tarifa por Hora (México):'
ws_total['A50'].font = Font(bold=True)

tarifas_headers = ['Nivel', 'Tarifa/Hora (MXN)', 'Tarifa/Hora (USD)', 'Horas', 'Subtotal (MXN)', 'Subtotal (USD)']
row = 51
for col, header in enumerate(tarifas_headers, start=1):
    cell = ws_total.cell(row, col, header)
    cell.fill = encabezado_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

tarifas_data = [
    ['Junior Developer', '$300-400', '$15-20', 200, '$70,000', '$3,500'],
    ['Mid Developer', '$500-700', '$25-35', 400, '$240,000', '$12,000'],
    ['Senior Developer', '$800-1,200', '$40-60', 128, '$128,000', '$6,400'],
    ['', '', '', 'TOTAL', '$438,000', '$21,900']
]

row = 52
for i, data in enumerate(tarifas_data):
    for col, value in enumerate(data, start=1):
        cell = ws_total.cell(row + i, col, value)
        cell.border = border
        if i == len(tarifas_data) - 1:
            cell.font = Font(bold=True, size=12)
            cell.fill = total_fill
            if col in [5, 6]:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in [2, 3, 4, 5, 6]:
            cell.alignment = Alignment(horizontal='center')

ws_total['A58'] = 'DURACIÓN ESTIMADA DEL PROYECTO'
ws_total['A58'].font = Font(bold=True, size=12, color="FFFFFF")
ws_total['A58'].fill = subtitulo_fill
ws_total.merge_cells('A58:F58')

ws_total['A59'] = 'Jornada de 8 horas/día:'
ws_total['B59'] = '728 horas ÷ 8 = 91 días'
ws_total['A60'] = 'Con equipo de 2 personas:'
ws_total['B60'] = '91 ÷ 2 = 46 días (≈ 2 meses)'
ws_total['B60'].font = Font(bold=True)
ws_total['A61'] = 'Con equipo de 3 personas:'
ws_total['B61'] = '91 ÷ 3 = 30 días (≈ 1 mes)'
ws_total['B61'].font = Font(bold=True)

ws_total['A64'] = 'RESUMEN EJECUTIVO DE COSTOS'
ws_total['A64'].font = Font(bold=True, size=12, color="FFFFFF")
ws_total['A64'].fill = titulo_fill
ws_total['A64'].alignment = Alignment(horizontal='center')
ws_total.merge_cells('A64:F64')

resumen_costos = [
    ['Concepto', 'Valor'],
    ['Puntos de Función Totales', '104 PF'],
    ['Líneas de Código Estimadas', '~2,600 LOC'],
    ['Horas de Desarrollo', '728 horas'],
    ['Costo Total (MXN)', '$438,000'],
    ['Costo Total (USD)', '$21,900'],
    ['Duración (1 persona)', '91 días'],
    ['Duración (equipo 2)', '46 días'],
    ['Duración (equipo 3)', '30 días']
]

row = 65
for i, data in enumerate(resumen_costos):
    for col, value in enumerate(data, start=1):
        cell = ws_total.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        else:
            if col == 1:
                cell.font = Font(bold=True)
            if col == 2:
                cell.alignment = Alignment(horizontal='center')

# Anchos de columna
ws_total.column_dimensions['A'].width = 35
ws_total.column_dimensions['B'].width = 20
ws_total.column_dimensions['C'].width = 60
ws_total.column_dimensions['D'].width = 15
ws_total.column_dimensions['E'].width = 20
ws_total.column_dimensions['F'].width = 20

# ============================================================================
# HOJA 8: METODOLOGÍA
# ============================================================================
ws_metod = wb.create_sheet("Metodología")

ws_metod['A1'] = 'METODOLOGÍA DE ANÁLISIS DE PUNTOS DE FUNCIÓN'
ws_metod['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_metod['A1'].fill = titulo_fill
ws_metod['A1'].alignment = Alignment(horizontal='center')
ws_metod.merge_cells('A1:D1')

metodologia_texto = """
ESTÁNDAR UTILIZADO: IFPUG (International Function Point Users Group) v4.3

1. COMPONENTES FUNCIONALES:

   a) Entradas Externas (EI):
      - Procesos que reciben datos desde fuera del sistema
      - Actualizan uno o más archivos lógicos internos
      - Valoración: Baja=3, Media=4, Alta=6 PF

   b) Salidas Externas (EO):
      - Envían datos procesados fuera del sistema
      - Incluyen lógica de cálculo o transformación
      - Valoración: Baja=4, Media=5, Alta=7 PF

   c) Consultas Externas (EQ):
      - Recuperan datos sin modificarlos
      - No incluyen lógica compleja de procesamiento
      - Valoración: Baja=3, Media=4, Alta=6 PF

   d) Archivos Lógicos Internos (ILF):
      - Grupos de datos mantenidos por el sistema
      - Valoración: Baja=7, Media=10, Alta=15 PF

   e) Archivos de Interfaz Externa (EIF):
      - Datos referenciados pero mantenidos por otros sistemas
      - Valoración: Baja=5, Media=7, Alta=10 PF

2. FACTOR DE AJUSTE (VAF):
   
   VAF = 0.65 + (0.01 × Suma de 14 Características)
   
   Cada característica se valora de 0 a 5:
   0 = Sin influencia
   1 = Influencia incidental
   2 = Influencia moderada
   3 = Influencia media
   4 = Influencia significativa
   5 = Influencia fuerte

3. CÁLCULO FINAL:
   
   UFP (Unadjusted Function Points) = Suma de todos los componentes
   AFP (Adjusted Function Points) = UFP × VAF

4. CONVERSIÓN A ESFUERZO:
   
   Horas = AFP × Factor de conversión (6-8 horas/PF promedio industria)

5. LIMITACIONES Y CONSIDERACIONES:
   
   - Esta es una estimación basada en análisis funcional
   - No incluye costos de infraestructura, licencias, etc.
   - Las tarifas pueden variar según región y experiencia
   - La duración puede variar según complejidad técnica y cambios de alcance
"""

ws_metod['A3'] = metodologia_texto
ws_metod.merge_cells('A3:D50')
ws_metod['A3'].alignment = Alignment(vertical='top', wrap_text=True)

ws_metod.column_dimensions['A'].width = 100

# Guardar archivo
wb.save('c:\\Users\\Joshua Rafael\\OneDrive\\Escritorio\\Asegurnasa\\ANALISIS_PUNTOS_FUNCION.xlsx')
print("✓ Archivo creado exitosamente: ANALISIS_PUNTOS_FUNCION.xlsx")
print("\nRESUMEN:")
print("- Puntos de Función sin ajustar (UFP): 98")
print("- Factor de Ajuste (VAF): 1.06")
print("- Puntos de Función Ajustados (AFP): 104 PF")
print("- Horas estimadas: 728 horas")
print("- Costo estimado: $438,000 MXN ($21,900 USD)")
print("- Duración estimada: 30-91 días según tamaño del equipo")
