"""
Cotización ajustada para cliente minorista/personal - ASEGURNASA
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cotización Minorista"

# Estilos
titulo_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
subtitulo_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
encabezado_fill = PatternFill(start_color="60a5fa", end_color="60a5fa", fill_type="solid")
precio_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
recomendado_fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezado
ws['A1'] = 'COTIZACIÓN - USO PERSONAL/MINORISTA'
ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws['A1'].fill = titulo_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:E1')
ws.row_dimensions[1].height = 30

ws['A2'] = 'Sistema de Gestión de Pólizas ASEGURNASA'
ws['A2'].font = Font(size=11, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.merge_cells('A2:E2')

ws['A3'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y")}'
ws['A3'].alignment = Alignment(horizontal='center')
ws.merge_cells('A3:E3')

# DIFERENCIA DE MERCADO
ws['A5'] = 'TIPO DE CLIENTE Y AJUSTE DE PRECIO'
ws['A5'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A5'].fill = subtitulo_fill
ws['A5'].alignment = Alignment(horizontal='center')
ws.merge_cells('A5:E5')

diferencia_data = [
    ['Tipo de Cliente', 'Características', 'Rango de Precio'],
    ['EMPRESARIAL', 'Empresa con múltiples usuarios, facturación', '$70,000 - $150,000 MXN'],
    ['PYME', 'Pequeño negocio, 2-5 usuarios', '$30,000 - $70,000 MXN'],
    ['MINORISTA/PERSONAL ⭐', 'Uso individual, gestión propia', '$8,000 - $20,000 MXN'],
]

row = 6
for i, data in enumerate(diferencia_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif '⭐' in str(value):
            cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
            cell.font = Font(bold=True)

ws.merge_cells('B6:B9')
ws.merge_cells('C6:C9')

# PRECIOS PARA MINORISTA
ws['A11'] = 'OPCIONES DE PRECIO PARA USO PERSONAL'
ws['A11'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A11'].fill = subtitulo_fill
ws['A11'].alignment = Alignment(horizontal='center')
ws.merge_cells('A11:E11')

precios_data = [
    ['Paquete', 'Incluye', 'MXN', 'USD', 'Observaciones'],
    ['CONOCIDO/AMIGO', 'Software + instalación básica', '$5,000', '$250', 'Precio de favor, sin soporte'],
    ['BÁSICO', 'Software + instalación + 1 capacitación', '$8,000', '$400', 'Sin soporte posterior'],
    ['ESTÁNDAR ⭐', 'Todo lo anterior + manual + 15 días soporte', '$12,000', '$600', 'Recomendado uso personal'],
    ['COMPLETO', 'Todo + 30 días soporte + actualizaciones', '$18,000', '$900', 'Máximo para minorista'],
]

row = 12
for i, data in enumerate(precios_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif '⭐' in str(value):
            cell.fill = recomendado_fill
            cell.font = Font(bold=True)
        
        if col in [3, 4] and i > 0:
            cell.alignment = Alignment(horizontal='center')

ws.merge_cells('E12:E16')

# PRECIO RECOMENDADO
ws['A18'] = '💰 PRECIO RECOMENDADO PARA TU CASO'
ws['A18'].font = Font(bold=True, size=13, color="FFFFFF")
ws['A18'].fill = titulo_fill
ws['A18'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A18:E18')
ws.row_dimensions[18].height = 25

ws['A19'] = 'Ya que es un CONOCIDO que solo quiere gestionar su información personal:'
ws['A19'].font = Font(italic=True)
ws['A19'].alignment = Alignment(horizontal='center')
ws.merge_cells('A19:E19')

precio_sugerido = [
    ['Modalidad', 'Precio MXN', 'Precio USD', 'Lo que incluye'],
    ['Precio Amigo/Favor', '$5,000 - $6,000', '$250 - $300', 'Software instalado, sin extras'],
    ['Precio Justo ⭐', '$10,000 - $12,000', '$500 - $600', 'Software + instalación + capacitación + manual'],
    ['Precio Máximo', '$15,000 - $18,000', '$750 - $900', 'Todo lo anterior + 30 días soporte'],
]

row = 20
for i, data in enumerate(precio_sugerido):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif '⭐' in str(value):
            cell.fill = recomendado_fill
            cell.font = Font(bold=True, size=12)
        
        if col in [2, 3] and i > 0:
            cell.alignment = Alignment(horizontal='center')

ws.merge_cells('D20:D23')

# JUSTIFICACIÓN DEL PRECIO
ws['A25'] = 'POR QUÉ ESTE RANGO DE PRECIO ES APROPIADO'
ws['A25'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A25'].fill = subtitulo_fill
ws.merge_cells('A25:E25')

justificacion = [
    ['Factor', 'Consideración'],
    ['✓ Cliente conocido', 'Puedes dar precio preferencial'],
    ['✓ Uso individual', 'No es para lucrar, solo gestión personal'],
    ['✓ Sin múltiples usuarios', 'No necesita licencias empresariales'],
    ['✓ Base instalada pequeña', 'Solo 1 computadora'],
    ['✓ Soporte limitado', 'No requiere soporte empresarial 24/7'],
    ['✓ Tu desarrollo ya hecho', 'No tienes que desarrollar desde cero'],
    ['✓ Mercado minorista', 'Menor capacidad de pago que empresas'],
]

row = 26
for i, data in enumerate(justificacion):
    col1 = ws.cell(row + i, 1, data[0])
    col2 = ws.cell(row + i, 3, data[1])
    col1.border = border
    col2.border = border
    if i == 0:
        col1.fill = encabezado_fill
        col2.fill = encabezado_fill
        col1.font = Font(bold=True, color="FFFFFF")
        col2.font = Font(bold=True, color="FFFFFF")
        col1.alignment = Alignment(horizontal='center')
        col2.alignment = Alignment(horizontal='center')
    else:
        col1.font = Font(bold=True, color="059669")
    
    ws.merge_cells(f'A{row + i}:B{row + i}')
    ws.merge_cells(f'C{row + i}:E{row + i}')

# COMPARACIÓN CON ALTERNATIVAS
ws['A35'] = 'COMPARACIÓN CON OTRAS OPCIONES DISPONIBLES'
ws['A35'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A35'].fill = subtitulo_fill
ws.merge_cells('A35:E35')

comparacion = [
    ['Alternativa', 'Costo', 'Ventaja', 'Desventaja'],
    ['Hoja de Excel manual', 'Gratis', 'Sin costo', 'Sin automatización, propenso a errores'],
    ['Google Sheets', 'Gratis', 'Colaborativo', 'Requiere internet, poco profesional'],
    ['Software SaaS (web)', '$500-1,500/mes', 'Actualizado', 'Costo recurrente ($6K-18K/año)'],
    ['Sistema genérico', '$3,000-5,000', 'Económico', 'No específico para seguros'],
    ['TU SISTEMA ⭐', '$10,000-12,000', 'Pago único, específico, profesional', 'Inversión inicial'],
]

row = 36
for i, data in enumerate(comparacion):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif '⭐' in str(value):
            cell.fill = recomendado_fill
            cell.font = Font(bold=True)

# RECOMENDACIÓN FINAL
ws['A43'] = '🎯 RECOMENDACIÓN FINAL'
ws['A43'].font = Font(bold=True, size=14, color="FFFFFF")
ws['A43'].fill = titulo_fill
ws['A43'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A43:E43')
ws.row_dimensions[43].height = 28

recomendacion = """
PRECIO SUGERIDO: $10,000 - $12,000 MXN ($500 - $600 USD)

Este precio es justo porque:

1. Es un precio accesible para uso personal/minorista
2. Refleja el valor real del software (104 PF de complejidad)
3. Es SIGNIFICATIVAMENTE más barato que suscripciones anuales
4. Es pago único (no mensualidades)
5. Le das un precio preferencial por ser conocido

FORMA DE PAGO SUGERIDA:
• Opción 1: Pago único de $10,000-12,000 al entregar
• Opción 2: $5,000 anticipo + $5,000-7,000 a 30 días
• Opción 3: 2 pagos mensuales de $5,000-6,000

LO QUE DEBES INCLUIR POR ESE PRECIO:
✓ Software instalado en su computadora
✓ Base de datos configurada
✓ 1-2 sesiones de capacitación (2 horas c/u)
✓ Manual de usuario en PDF
✓ 15 días de soporte vía WhatsApp/teléfono
✓ Corrección de errores inmediatos

LO QUE NO DEBES INCLUIR (cobrar aparte si lo pide):
✗ Instalación en múltiples computadoras
✗ Soporte más allá de 15 días
✗ Nuevas funcionalidades o personalizaciones
✗ Capacitación adicional (cobra $500-1,000/sesión)
✗ Actualizaciones mayores (cobra $1,500-3,000)

IMPORTANTE:
- NO vendas por menos de $8,000 (estarías regalando tu trabajo)
- NO incluyas soporte ilimitado (que pague soporte mensual si lo necesita)
- SÍ explica que es pago único vs suscripciones
- SÍ enfatiza que es específico para seguros de vehículos
"""

ws['A44'] = recomendacion
ws.merge_cells('A44:E75')
ws['A44'].alignment = Alignment(vertical='top', wrap_text=True)
ws['A44'].font = Font(size=10)

# PLANTILLA DE MENSAJE
ws['A77'] = 'PLANTILLA DE MENSAJE PARA TU CONOCIDO'
ws['A77'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A77'].fill = subtitulo_fill
ws.merge_cells('A77:E77')

mensaje = """
Hola [Nombre],

Te tengo lista la solución para gestionar tus pólizas de seguros. Desarrollé un sistema 
profesional específicamente diseñado para administración de pólizas de vehículos.

🖥️ ¿QUÉ INCLUYE?
✓ Sistema completo para gestionar todas tus pólizas
✓ Búsqueda instantánea por cualquier dato
✓ Exportación a Excel y PDF profesionales
✓ Cálculo automático de vencimientos con alertas
✓ Catálogos completos de estados y municipios de México
✓ Interfaz moderna y muy fácil de usar

📦 PAQUETE COMPLETO:
• Software instalado en tu computadora
• Configuración personalizada
• 2 sesiones de capacitación (te enseño a usarlo)
• Manual de usuario
• 15 días de soporte por WhatsApp

💰 INVERSIÓN: $10,000 pesos (pago único)
    Precio especial para ti (normalmente $75,000 para empresas)

💳 FORMA DE PAGO:
    Opción 1: $10,000 al entregar
    Opción 2: $5,000 ahora + $5,000 en 30 días

⏱️ ENTREGA: Inmediata (ya está listo)

IMPORTANTE: Este es un pago único, NO tienes que pagar mensualidades como 
con sistemas en línea que te cobran $500-1,500 cada mes.

¿Te parece bien? Puedo instalártelo cuando gustes.

Saludos,
[Tu nombre]
"""

ws['A78'] = mensaje
ws.merge_cells('A78:E105')
ws['A78'].alignment = Alignment(vertical='top', wrap_text=True)
ws['A78'].border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# OPCIONES SI REGATEA
ws['A107'] = 'SI TU CONOCIDO REGATEA O DICE QUE ES MUCHO'
ws['A107'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A107'].fill = subtitulo_fill
ws.merge_cells('A107:E107')

regateo_data = [
    ['Si dice...', 'Tú respondes...', 'Acción'],
    ['"Es mucho dinero"', 'Compara: sistemas web cuestan $500/mes = $6,000/año', 'Muestra que se recupera en 2 años'],
    ['"¿No hay más barato?"', 'Puedo darte solo el software por $8,000 sin capacitación ni soporte', 'Ofrece versión reducida'],
    ['"¿Me lo regalas?"', 'Son 104 puntos de función, ~700 horas de desarrollo', 'Explica el trabajo invertido'],
    ['"¿Y si te pago...?"', 'El mínimo es $8,000, ya es precio de amigo', 'Mantén tu precio mínimo'],
]

row = 108
for i, data in enumerate(regateo_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        if col == 2:
            cell.font = Font(bold=True, color="059669")

# RESUMEN VISUAL
ws['A114'] = '📊 RESUMEN EJECUTIVO'
ws['A114'].font = Font(bold=True, size=13, color="FFFFFF")
ws['A114'].fill = titulo_fill
ws['A114'].alignment = Alignment(horizontal='center')
ws.merge_cells('A114:E114')

resumen_final = [
    ['Concepto', 'Valor'],
    ['PRECIO MÍNIMO ABSOLUTO', '$8,000 MXN ($400 USD)'],
    ['⭐ PRECIO RECOMENDADO', '$10,000 - $12,000 MXN ($500-600 USD)'],
    ['PRECIO MÁXIMO MINORISTA', '$18,000 MXN ($900 USD)'],
    ['', ''],
    ['Forma de pago sugerida', '$5,000 anticipo + $5,000-7,000 a 30 días'],
    ['Tiempo de entrega', 'Inmediato'],
    ['Soporte incluido', '15 días'],
    ['Tipo de cliente', 'Uso personal/minorista'],
]

row = 115
for i, data in enumerate(resumen_final):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if '⭐' in str(value):
            cell.fill = recomendado_fill
            cell.font = Font(bold=True, size=12)
        if col == 1 and i > 0:
            cell.font = Font(bold=True)
        if col == 2 and i > 0:
            cell.alignment = Alignment(horizontal='center')

# Ajustar columnas
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 30

# Guardar
wb.save('c:\\Users\\Joshua Rafael\\OneDrive\\Escritorio\\Asegurnasa\\COTIZACION_MINORISTA.xlsx')

print("=" * 80)
print("✓ COTIZACIÓN MINORISTA CREADA EXITOSAMENTE")
print("=" * 80)
print("\n💰 PRECIOS AJUSTADOS PARA USO PERSONAL/MINORISTA:\n")
print("   Precio Mínimo:       $8,000 MXN   ($400 USD)")
print("   ⭐ RECOMENDADO:       $10,000-12,000 MXN ($500-600 USD)")
print("   Precio Máximo:       $18,000 MXN  ($900 USD)")
print("\n" + "=" * 80)
print("\n🎯 RECOMENDACIÓN PARA TU CONOCIDO:")
print("   • Cobra $10,000 - $12,000 MXN (pago único)")
print("   • Incluye: software + instalación + capacitación + manual + 15 días soporte")
print("   • Forma de pago: $5,000 anticipo + resto a 30 días")
print("   • NO vendas por menos de $8,000 (estarías regalando tu trabajo)")
print("\n" + "=" * 80)
print("\n💡 ARGUMENTO CLAVE:")
print('   "Los sistemas web cuestan $500-1,500/mes ($6K-18K/año)')
print('    Mi sistema es pago único de $10K-12K, sin mensualidades"')
print("\n" + "=" * 80)
print("\n📄 Archivo creado: COTIZACION_MINORISTA.xlsx")
print("=" * 80)
