"""
Generador de Cotización Profesional para Proyecto ASEGURNASA
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cotización Proyecto"

# Estilos
titulo_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
subtitulo_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
encabezado_fill = PatternFill(start_color="60a5fa", end_color="60a5fa", fill_type="solid")
precio_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
recomendado_fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezado
ws['A1'] = 'COTIZACIÓN PROFESIONAL'
ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
ws['A1'].fill = titulo_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:F1')
ws.row_dimensions[1].height = 35

ws['A2'] = 'Sistema de Gestión de Pólizas de Seguros - ASEGURNASA'
ws['A2'].font = Font(size=12, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.merge_cells('A2:F2')

ws['A3'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y")}'
ws['A3'].alignment = Alignment(horizontal='center')
ws.merge_cells('A3:F3')

# ANÁLISIS DEL PROYECTO
ws['A5'] = 'ANÁLISIS DEL PROYECTO DESARROLLADO'
ws['A5'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A5'].fill = subtitulo_fill
ws['A5'].alignment = Alignment(horizontal='center')
ws.merge_cells('A5:F5')

analisis_data = [
    ['Métrica', 'Valor', 'Observación'],
    ['Puntos de Función', '104 PF', 'Proyecto de complejidad media-alta'],
    ['Líneas de Código', '~2,600 LOC', 'Python profesional'],
    ['Módulos desarrollados', '3 archivos', 'main.py, database.py, datos_mexico.py'],
    ['Funcionalidades principales', '17 funciones', 'CRUD completo + exportación + búsquedas'],
    ['Tablas de base de datos', '1 tabla', '31 campos con relaciones complejas'],
    ['Interfaces de usuario', '1 GUI completa', 'Diseño profesional con customtkinter'],
    ['Catálogos implementados', '500+ registros', '32 estados + municipios de México'],
    ['Formatos de exportación', '2 (Excel + PDF)', 'Con formato profesional']
]

row = 6
for i, data in enumerate(analisis_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        else:
            if col == 1:
                cell.font = Font(bold=True)

# MODALIDADES DE COBRO
ws['A17'] = 'MODALIDADES DE COBRO RECOMENDADAS'
ws['A17'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A17'].fill = subtitulo_fill
ws['A17'].alignment = Alignment(horizontal='center')
ws.merge_cells('A17:F17')

ws['A18'] = 'Considera que como desarrollador independiente en México, puedes cobrar según diferentes modelos:'
ws['A18'].font = Font(italic=True)
ws.merge_cells('A18:F18')

# OPCIÓN 1: POR PUNTOS DE FUNCIÓN
ws['A20'] = 'OPCIÓN 1: PRECIO POR PUNTOS DE FUNCIÓN'
ws['A20'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A20'].fill = encabezado_fill
ws.merge_cells('A20:F20')

pf_data = [
    ['Nivel', 'Precio/PF (MXN)', 'Total (104 PF)', 'Total (USD)', 'Descripción'],
    ['Freelancer Junior', '$300 - $500', '$31,200 - $52,000', '$1,560 - $2,600', 'Desarrollador con 0-2 años experiencia'],
    ['Freelancer Mid', '$600 - $900', '$62,400 - $93,600', '$3,120 - $4,680', 'Desarrollador con 2-5 años experiencia'],
    ['Freelancer Senior', '$1,000 - $1,500', '$104,000 - $156,000', '$5,200 - $7,800', 'Desarrollador con 5+ años experiencia'],
    ['', '', 'RANGO RECOMENDADO:', '$62,400 - $104,000', '$3,120 - $5,200']
]

row = 21
for i, data in enumerate(pf_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif i == len(pf_data) - 1:
            cell.fill = recomendado_fill
            cell.font = Font(bold=True, color="FFFFFF")
        if col in [2, 3, 4]:
            cell.alignment = Alignment(horizontal='center')

# OPCIÓN 2: PRECIO FIJO POR PROYECTO
ws['A28'] = 'OPCIÓN 2: PRECIO FIJO POR PROYECTO COMPLETO'
ws['A28'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A28'].fill = encabezado_fill
ws.merge_cells('A28:F28')

fijo_data = [
    ['Paquete', 'Incluye', 'Precio MXN', 'Precio USD', 'Tiempo Entrega'],
    ['BÁSICO', 'Sistema + instalación + 1 capacitación', '$45,000', '$2,250', '5 días soporte'],
    ['PROFESIONAL ⭐', 'Sistema + instalación + 2 capacitaciones + manual + 30 días soporte', '$75,000', '$3,750', '30 días soporte'],
    ['EMPRESARIAL', 'Todo lo anterior + personalización logo + 90 días soporte + actualizaciones', '$110,000', '$5,500', '90 días soporte'],
    ['', '', 'RECOMENDADO:', '$75,000', '$3,750']
]

row = 29
for i, data in enumerate(fijo_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif '⭐' in str(data[0]):
            cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
            cell.font = Font(bold=True)
        elif i == len(fijo_data) - 1:
            cell.fill = recomendado_fill
            cell.font = Font(bold=True, color="FFFFFF")
        if col in [3, 4]:
            cell.alignment = Alignment(horizontal='center')

# OPCIÓN 3: POR HORAS TRABAJADAS
ws['A35'] = 'OPCIÓN 3: COBRO POR HORAS REALMENTE TRABAJADAS'
ws['A35'].font = Font(bold=True, size=11, color="FFFFFF")
ws['A35'].fill = encabezado_fill
ws.merge_cells('A35:F35')

ws['A36'] = 'Si llevaste registro de horas reales invertidas:'
ws['A36'].font = Font(italic=True)
ws.merge_cells('A36:F36')

horas_data = [
    ['Estimación de Horas', 'Tarifa/Hora MXN', 'Tarifa/Hora USD', 'Total MXN', 'Total USD'],
    ['40 horas (1 semana intensiva)', '$500 - $800', '$25 - $40', '$20,000 - $32,000', '$1,000 - $1,600'],
    ['80 horas (2 semanas)', '$500 - $800', '$25 - $40', '$40,000 - $64,000', '$2,000 - $3,200'],
    ['120 horas (3 semanas)', '$500 - $800', '$25 - $40', '$60,000 - $96,000', '$3,000 - $4,800'],
    ['160 horas (1 mes)', '$500 - $800', '$25 - $40', '$80,000 - $128,000', '$4,000 - $6,400'],
]

row = 37
for i, data in enumerate(horas_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        if col in [2, 3, 4, 5]:
            cell.alignment = Alignment(horizontal='center')

# RECOMENDACIÓN FINAL
ws['A44'] = 'RECOMENDACIÓN DE PRECIO ÓPTIMO'
ws['A44'].font = Font(bold=True, size=13, color="FFFFFF")
ws['A44'].fill = titulo_fill
ws['A44'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A44:F44')
ws.row_dimensions[44].height = 30

recomendacion_data = [
    ['Concepto', 'Rango Mínimo', 'Rango Óptimo', 'Rango Premium'],
    ['Precio Base del Sistema', '$45,000', '$75,000', '$110,000'],
    ['+ Instalación y Configuración', '+$3,000', '+$5,000', '+$8,000'],
    ['+ Capacitación (por sesión)', '+$2,000/sesión', '+$3,000/sesión', '+$5,000/sesión'],
    ['+ Manual de Usuario', '+$3,000', '+$5,000', '+$8,000'],
    ['+ Soporte Técnico (mensual)', '+$2,000/mes', '+$4,000/mes', '+$6,000/mes'],
    ['', '', '', ''],
    ['TOTAL RECOMENDADO', '$50,000 - $55,000', '$82,000 - $92,000', '$130,000 - $145,000'],
    ['En USD', '$2,500 - $2,750', '$4,100 - $4,600', '$6,500 - $7,250']
]

row = 45
for i, data in enumerate(recomendacion_data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif i == len(recomendacion_data) - 2:
            cell.fill = precio_fill
            cell.font = Font(bold=True, size=12, color="FFFFFF")
        elif i == len(recomendacion_data) - 1:
            cell.fill = precio_fill
            cell.font = Font(bold=True, size=11, color="FFFFFF")
        
        if col == 1 and i > 0 and i < len(recomendacion_data) - 2:
            cell.font = Font(bold=True)
        if col > 1:
            cell.alignment = Alignment(horizontal='center')

# PRECIO SUGERIDO
ws['A55'] = '💰 PRECIO SUGERIDO PARA COBRAR'
ws['A55'].font = Font(bold=True, size=14, color="000000")
ws['A55'].fill = PatternFill(start_color="fbbf24", end_color="fbbf24", fill_type="solid")
ws['A55'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A55:F55')
ws.row_dimensions[55].height = 30

ws['A56'] = 'Considerando la complejidad del sistema (104 PF), diseño profesional y funcionalidad completa:'
ws['A56'].font = Font(italic=True)
ws['A56'].alignment = Alignment(horizontal='center')
ws.merge_cells('A56:F56')

precio_sugerido = [
    ['', 'MXN', 'USD'],
    ['Precio Mínimo Aceptable', '$50,000', '$2,500'],
    ['Precio Justo y Competitivo ⭐', '$75,000', '$3,750'],
    ['Precio Premium (con soporte)', '$95,000', '$4,750']
]

row = 57
for i, data in enumerate(precio_sugerido):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row + i, col, value)
        cell.border = border
        if i == 0:
            cell.fill = encabezado_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        elif '⭐' in str(data[0]):
            cell.fill = recomendado_fill
            cell.font = Font(bold=True, size=13, color="FFFFFF")
        
        if col == 1 and i > 0 and '⭐' not in str(data[0]):
            cell.font = Font(bold=True)
        if col > 1:
            cell.alignment = Alignment(horizontal='center')
            if i > 0 and i < len(precio_sugerido) - 1:
                cell.font = Font(bold=True, size=12)

# FACTORES QUE JUSTIFICAN EL PRECIO
ws['A62'] = 'FACTORES QUE JUSTIFICAN EL PRECIO'
ws['A62'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A62'].fill = subtitulo_fill
ws.merge_cells('A62:F62')

justificacion_data = [
    ['✓ Interface gráfica profesional moderna (customtkinter)', 'No es un simple formulario'],
    ['✓ Base de datos SQLite con 31 campos optimizados', 'Estructura compleja y escalable'],
    ['✓ Sistema CRUD completo', 'Crear, Leer, Actualizar, Eliminar'],
    ['✓ Búsqueda en tiempo real multi-campo', 'Funcionalidad avanzada'],
    ['✓ Exportación profesional a Excel y PDF', 'Con formato y diseño'],
    ['✓ Catálogo completo de México (32 estados + municipios)', '500+ registros geográficos'],
    ['✓ Tabla dinámica de coberturas con JSON', 'Almacenamiento flexible'],
    ['✓ Cálculos automáticos de vencimientos', 'Lógica de negocio compleja'],
    ['✓ Sistema de alertas visuales con colores', 'UX profesional'],
    ['✓ Validaciones de datos en tiempo real', 'Control de calidad'],
    ['✓ Diseño responsive y escalable', 'Adaptable a diferentes pantallas'],
    ['✓ Código limpio y documentado', 'Fácil mantenimiento futuro']
]

row = 63
for i, data in enumerate(justificacion_data):
    cell1 = ws.cell(row + i, 1, data[0])
    cell2 = ws.cell(row + i, 4, data[1])
    cell1.border = border
    cell2.border = border
    cell1.font = Font(bold=True, color="059669")
    cell2.font = Font(italic=True)
    ws.merge_cells(f'A{row + i}:C{row + i}')
    ws.merge_cells(f'D{row + i}:F{row + i}')

# CONSEJOS DE NEGOCIACIÓN
ws['A77'] = 'CONSEJOS PARA CERRAR LA VENTA'
ws['A77'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A77'].fill = subtitulo_fill
ws.merge_cells('A77:F77')

consejos = """
1. Presenta el valor, no solo el precio: Enfatiza las 17 funcionalidades y la complejidad del sistema

2. Ofrece paquetes escalonados: Básico ($50K), Profesional ($75K), Premium ($95K)

3. Incluye servicios adicionales:
   - Instalación y configuración personalizada
   - 2-3 sesiones de capacitación presencial o remota
   - Manual de usuario digital en PDF
   - 30 días de soporte técnico incluido
   - 1 revisión/actualización menor gratis

4. Forma de pago flexible:
   - 50% al inicio (anticipo)
   - 50% al terminar (contra entrega)
   O bien: 40% inicio, 30% avance al 50%, 30% entrega final

5. Garantía de software:
   - 30 días para corrección de errores sin costo
   - Garantía de funcionamiento según especificaciones

6. Compara con alternativas:
   - Sistemas web con suscripción mensual: $1,000-3,000/mes ($12K-36K/año)
   - Tu solución: Pago único sin mensualidades
   - Software local: Más rápido, no depende de internet

7. Muestra el análisis de Puntos de Función:
   - 104 PF × $720/PF promedio = $75,000 (respaldo técnico del precio)
"""

ws['A78'] = consejos
ws.merge_cells('A78:F95')
ws['A78'].alignment = Alignment(vertical='top', wrap_text=True)

# PLANTILLA DE PROPUESTA
ws['A97'] = 'PLANTILLA DE CORREO PARA ENVIAR AL CLIENTE'
ws['A97'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A97'].fill = titulo_fill
ws.merge_cells('A97:F97')

plantilla = """
Asunto: Propuesta - Sistema de Gestión de Pólizas ASEGURNASA

Estimado/a [Nombre del Cliente],

Me complace presentarle el "Sistema de Gestión de Pólizas de Seguros ASEGURNASA", una solución profesional desarrollada específicamente para optimizar la administración de pólizas de vehículos.

CARACTERÍSTICAS PRINCIPALES:
✓ Gestión completa de pólizas (crear, editar, eliminar, buscar)
✓ Interface gráfica profesional y fácil de usar
✓ Búsqueda instantánea por cualquier campo
✓ Exportación a Excel y PDF con formato profesional
✓ Catálogos completos de estados y municipios de México
✓ Cálculo automático de vencimientos con alertas visuales
✓ Tabla dinámica de coberturas contratadas
✓ Sistema 100% local (no requiere internet, datos seguros)

INVERSIÓN:
Paquete Profesional: $75,000 MXN ($3,750 USD)

INCLUYE:
• Software completo instalado en su equipo
• Instalación y configuración personalizada
• 2 sesiones de capacitación para su equipo
• Manual de usuario digital
• 30 días de soporte técnico incluido
• Garantía de 30 días por errores de funcionamiento

FORMA DE PAGO:
• 50% anticipo ($37,500)
• 50% contra entrega ($37,500)

TIEMPO DE ENTREGA: Inmediato (sistema ya desarrollado)

Este sistema ha sido diseñado con 104 Puntos de Función, lo que representa aproximadamente 728 horas de desarrollo profesional. El precio ofrecido es altamente competitivo considerando la inversión en desarrollo y las funcionalidades incluidas.

Quedo a sus órdenes para cualquier duda o demostración del sistema.

Saludos cordiales,
[Tu Nombre]
[Tu Teléfono]
[Tu Email]
"""

ws['A98'] = plantilla
ws.merge_cells('A98:F125')
ws['A98'].alignment = Alignment(vertical='top', wrap_text=True)
ws['A98'].border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20

# Guardar
wb.save('c:\\Users\\Joshua Rafael\\OneDrive\\Escritorio\\Asegurnasa\\COTIZACION_PROFESIONAL.xlsx')
print("=" * 70)
print("✓ COTIZACIÓN CREADA EXITOSAMENTE")
print("=" * 70)
print("\n📊 RESUMEN DE PRECIOS RECOMENDADOS:\n")
print("   Mínimo Aceptable:    $50,000 MXN  ($2,500 USD)")
print("   ⭐ RECOMENDADO:       $75,000 MXN  ($3,750 USD)")
print("   Premium:             $95,000 MXN  ($4,750 USD)")
print("\n" + "=" * 70)
print("\n💡 CONSEJOS CLAVE:")
print("   • Enfatiza el VALOR, no solo el precio")
print("   • Ofrece paquetes (Básico/Profesional/Premium)")
print("   • Incluye capacitación y soporte")
print("   • Pide 50% anticipo, 50% contra entrega")
print("   • Compara con sistemas de suscripción mensual")
print("\n" + "=" * 70)
print("\n📄 Archivo creado: COTIZACION_PROFESIONAL.xlsx")
print("=" * 70)
