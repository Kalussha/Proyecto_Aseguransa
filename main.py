"""
Sistema de Gestión de Pólizas de Seguros de Vehículos
Aplicación de escritorio moderna con customtkinter y SQLite.
"""
import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime, timedelta
import openpyxl
import json
import requests
import threading
from database import DatabaseManager
from datos_mexico import (
    ESTADOS_MEXICO, MUNICIPIOS_POR_ESTADO, TIPOS_MONEDA,
    USOS_VEHICULO, SERVICIOS_VEHICULO, FORMAS_PAGO,
    MESES_POR_FORMA_PAGO
)
import re


# Tema de colores Premium Moderno
COLORES = {
    "primario": "#2563eb",       # Azul vibrante moderno
    "primario_hover": "#1d4ed8",
    "secundario": "#3f3f46",     # Gris medio oscuro
    "exito": "#10b981",          # Verde esmeralda (éxito)
    "exito_hover": "#059669",
    "peligro": "#e11d48",        # Rojo rosado (peligro)
    "peligro_hover": "#be123c",
    "advertencia": "#f59e0b",    # Ámbar
    "info": "#0ea5e9",           # Azul cielo
    "fondo": "#09090b",          # Fondo principal casi negro (zinc-950)
    "fondo_claro": "#18181b",    # Paneles secundarios (zinc-900)
    "texto": "#f4f4f5",          # Texto principal claro
    "texto_secundario": "#a1a1aa", # Texto tenue
    "borde": "#27272a"           # Bordes sutiles
}


class CambiarPasswordWindow(ctk.CTkToplevel):
    """Ventana para cambiar contraseña (obligatoria o voluntaria)."""
    
    def __init__(self, parent, db, usuario_data, obligatorio=False):
        super().__init__(parent)
        
        self.db = db
        self.usuario_data = usuario_data
        self.obligatorio = obligatorio
        self.password_cambiado = False
        
        # Configuración de ventana
        titulo = "Cambio de Contraseña Obligatorio" if obligatorio else "Cambiar Contraseña"
        self.title(f"ASEGURANZA - {titulo}")
        self.geometry("500x600")
        self.resizable(False, False)
        
        # Centrar ventana
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        # Hacer modal
        self.transient(parent)
        self.grab_set()
        
        # Si es obligatorio, no permitir cerrar
        if obligatorio:
            self.protocol("WM_DELETE_WINDOW", self.no_cerrar)
        
        self.crear_interfaz()
        
    def no_cerrar(self):
        """Previene cerrar la ventana si es cambio obligatorio."""
        messagebox.showwarning(
            "Cambio Obligatorio",
            "Debe cambiar su contraseña antes de continuar",
            parent=self
        )
    
    def crear_interfaz(self):
        """Crea la interfaz de cambio de contraseña."""
        main_frame = ctk.CTkFrame(self, fg_color=COLORES["fondo_claro"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Encabezado
        if self.obligatorio:
            header_frame = ctk.CTkFrame(main_frame, fg_color=COLORES["peligro"],)
            icono_texto = "⚠️"
            titulo_texto = "CAMBIO DE CONTRASEÑA REQUERIDO"
            mensaje = "Por seguridad, debe cambiar su contraseña antes de continuar usando el sistema."
        else:
            header_frame = ctk.CTkFrame(main_frame, fg_color=COLORES["primario"],)
            icono_texto = "🔒"
            titulo_texto = "CAMBIAR CONTRASEÑA"
            mensaje = "Ingrese su contraseña actual y defina una nueva contraseña segura."
        
        header_frame.pack(fill="x", pady=(0, 20))
        
        titulo = ctk.CTkLabel(
            header_frame,
            text=f"{icono_texto} {titulo_texto}",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="white"
        )
        titulo.pack(pady=15)
        
        # Mensaje informativo
        info_label = ctk.CTkLabel(
            main_frame,
            text=mensaje,
            font=ctk.CTkFont(size=12),
            text_color=COLORES["texto_secundario"],
            wraplength=450
        )
        info_label.pack(pady=(0, 20))
        
        # Usuario actual
        usuario_frame = ctk.CTkFrame(main_frame, fg_color=COLORES["info"],)
        usuario_frame.pack(fill="x", pady=(0, 20), padx=10)
        
        usuario_label = ctk.CTkLabel(
            usuario_frame,
            text=f"👤 Usuario: {self.usuario_data.get('usuario', 'N/A')}",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="white"
        )
        usuario_label.pack(pady=10)
        
        # Formulario
        form_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        form_frame.pack(fill="both", expand=True, padx=10)
        
        # Contraseña actual
        actual_label = ctk.CTkLabel(
            form_frame,
            text="Contraseña Actual:",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        actual_label.pack(fill="x", pady=(5, 5))
        
        self.actual_entry = ctk.CTkEntry(
            form_frame,
            placeholder_text="Ingrese su contraseña actual",
            show="●",
            height=40,
            font=ctk.CTkFont(size=13,),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.actual_entry.pack(fill="x", pady=(0, 15))
        self.actual_entry.bind('<Return>', lambda e: self.nueva_entry.focus())
        
        # Nueva contraseña
        nueva_label = ctk.CTkLabel(
            form_frame,
            text="Nueva Contraseña:",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        nueva_label.pack(fill="x", pady=(5, 5))
        
        self.nueva_entry = ctk.CTkEntry(
            form_frame,
            placeholder_text="Ingrese su nueva contraseña",
            show="●",
            height=40,
            font=ctk.CTkFont(size=13,),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.nueva_entry.pack(fill="x", pady=(0, 15))
        self.nueva_entry.bind('<Return>', lambda e: self.confirmar_entry.focus())
        
        # Confirmar nueva contraseña
        confirmar_label = ctk.CTkLabel(
            form_frame,
            text="Confirmar Nueva Contraseña:",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        confirmar_label.pack(fill="x", pady=(5, 5))
        
        self.confirmar_entry = ctk.CTkEntry(
            form_frame,
            placeholder_text="Confirme su nueva contraseña",
            show="●",
            height=40,
            font=ctk.CTkFont(size=13,),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.confirmar_entry.pack(fill="x", pady=(0, 15))
        self.confirmar_entry.bind('<Return>', lambda e: self.cambiar_password())
        
        # Requisitos de contraseña
        requisitos_frame = ctk.CTkFrame(form_frame, fg_color=COLORES["fondo_claro"],)
        requisitos_frame.pack(fill="x", pady=(10, 20))
        
        requisitos_label = ctk.CTkLabel(
            requisitos_frame,
            text="📋 Requisitos de contraseña:\n• Mínimo 8 caracteres\n• Se recomienda incluir letras, números y símbolos",
            font=ctk.CTkFont(size=11),
            text_color=COLORES["texto_secundario"],
            justify="left"
        )
        requisitos_label.pack(pady=10, padx=10)
        
        # Botones
        botones_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        botones_frame.pack(fill="x", padx=10, pady=(10, 0))
        
        self.cambiar_btn = ctk.CTkButton(
            botones_frame,
            text="Cambiar Contraseña",
            command=self.cambiar_password,
            height=45,
            font=ctk.CTkFont(size=14, weight="bold",),
            fg_color=COLORES["exito"],
            hover_color=COLORES["exito_hover"]
        )
        self.cambiar_btn.pack(fill="x", pady=(0, 10))
        
        if not self.obligatorio:
            cancelar_btn = ctk.CTkButton(
                botones_frame,
                text="Cancelar",
                command=self.destroy,
                height=40,
                font=ctk.CTkFont(size=13,),
                fg_color=COLORES["borde"],
                hover_color=COLORES["texto_secundario"]
            )
            cancelar_btn.pack(fill="x")
        
        # Focus inicial
        self.actual_entry.focus()
    
    def cambiar_password(self):
        """Procesa el cambio de contraseña."""
        actual = self.actual_entry.get()
        nueva = self.nueva_entry.get()
        confirmar = self.confirmar_entry.get()
        
        # Validaciones
        if not actual or not nueva or not confirmar:
            messagebox.showerror(
                "Error",
                "Todos los campos son obligatorios",
                parent=self
            )
            return
        
        if len(nueva) < 8:
            messagebox.showerror(
                "Error",
                "La nueva contraseña debe tener al menos 8 caracteres",
                parent=self
            )
            return
        
        if nueva != confirmar:
            messagebox.showerror(
                "Error",
                "La nueva contraseña y su confirmación no coinciden",
                parent=self
            )
            self.confirmar_entry.delete(0, 'end')
            self.confirmar_entry.focus()
            return
        
        # Intentar cambiar contraseña
        exito, mensaje = self.db.cambiar_password(
            self.usuario_data['usuario'],
            actual,
            nueva
        )
        
        if exito:
            self.password_cambiado = True
            messagebox.showinfo(
                "Éxito",
                mensaje,
                parent=self
            )
            self.destroy()
        else:
            messagebox.showerror(
                "Error",
                mensaje,
                parent=self
            )
            self.actual_entry.delete(0, 'end')
            self.actual_entry.focus()


class GestionUsuariosWindow(ctk.CTkToplevel):
    """Ventana para gestionar usuarios del sistema."""
    
    def __init__(self, parent, db):
        super().__init__(parent)
        
        self.db = db
        
        # Configuración de ventana
        self.title("ASEGURANZA - Gestión de Usuarios")
        self.geometry("900x700")
        
        # Centrar ventana
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        # Variables para SMTP
        self.smtp_config = {
            'servidor': 'smtp.gmail.com',
            'puerto': 587,
            'email': '',
            'password': ''
        }
        
        self.crear_interfaz()
        self.cargar_usuarios()
    
    def crear_interfaz(self):
        """Crea la interfaz de gestión de usuarios."""
        # Frame principal
        main_frame = ctk.CTkFrame(self, fg_color=COLORES["fondo_claro"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        header_frame = ctk.CTkFrame(main_frame, fg_color=COLORES["primario"],)
        header_frame.pack(fill="x", pady=(0, 20))
        
        titulo = ctk.CTkLabel(
            header_frame,
            text="👥 GESTIÓN DE USUARIOS",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        )
        titulo.pack(pady=15)
        
        # Botón crear usuario
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 15))
        
        crear_btn = ctk.CTkButton(
            btn_frame,
            text="➕ Crear Nuevo Usuario",
            command=self.abrir_crear_usuario,
            height=40,
            font=ctk.CTkFont(size=14, weight="bold",),
            fg_color=COLORES["exito"],
            hover_color=COLORES["exito_hover"]
        )
        crear_btn.pack(side="left", padx=5)
        
        # Lista de usuarios
        self.usuarios_frame = ctk.CTkScrollableFrame(
            main_frame,
            fg_color=COLORES["fondo"],
            
        )
        self.usuarios_frame.pack(fill="both", expand=True)
    
    def cargar_usuarios(self):
        """Carga la lista de usuarios."""
        # Limpiar frame
        for widget in self.usuarios_frame.winfo_children():
            widget.destroy()
        
        usuarios = self.db.listar_usuarios()
        
        for usuario in usuarios:
            self.crear_card_usuario(usuario)
    
    def crear_card_usuario(self, usuario):
        """Crea una tarjeta para un usuario."""
        card = ctk.CTkFrame(
            self.usuarios_frame,
            fg_color=COLORES["fondo_claro"],
            border_width=2,
            border_color=COLORES["borde"]
        )
        card.pack(fill="x", padx=5, pady=5)
        
        # Contenido
        content_frame = ctk.CTkFrame(card, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=15, pady=10)
        
        # Izquierda: Info usuario
        info_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        info_frame.pack(side="left", fill="both", expand=True)
        
        # Nombre y usuario
        nombre_label = ctk.CTkLabel(
            info_frame,
            text=f"👤 {usuario.get('nombre_completo', 'Sin nombre')}",
            font=ctk.CTkFont(size=14, weight="bold"),
            anchor="w"
        )
        nombre_label.pack(anchor="w")
        
        usuario_label = ctk.CTkLabel(
            info_frame,
            text=f"Usuario: {usuario.get('usuario', 'N/A')}",
            font=ctk.CTkFont(size=12),
            text_color=COLORES["texto_secundario"],
            anchor="w"
        )
        usuario_label.pack(anchor="w", pady=(2, 0))
        
        if usuario.get('email'):
            email_label = ctk.CTkLabel(
                info_frame,
                text=f"📧 {usuario.get('email')}",
                font=ctk.CTkFont(size=11),
                text_color=COLORES["texto_secundario"],
                anchor="w"
            )
            email_label.pack(anchor="w", pady=(2, 0))
        
        # Badges de estado
        badges_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
        badges_frame.pack(anchor="w", pady=(5, 0))
        
        if usuario.get('activo') == 1:
            estado_badge = ctk.CTkLabel(
                badges_frame,
                text="✓ Activo",
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color="white",
                fg_color=COLORES["exito"],
                corner_radius=4,
                padx=8,
                pady=2
            )
            estado_badge.pack(side="left", padx=(0, 5))
        else:
            estado_badge = ctk.CTkLabel(
                badges_frame,
                text="✗ Inactivo",
                font=ctk.CTkFont(size=10, weight="bold"),
                text_color="white",
                fg_color=COLORES["peligro"],
                corner_radius=4,
                padx=8,
                pady=2
            )
            estado_badge.pack(side="left", padx=(0, 5))
        
        if usuario.get('cambiar_password') == 1:
            pass_badge = ctk.CTkLabel(
                badges_frame,
                text="⚠️ Debe cambiar contraseña",
                font=ctk.CTkFont(size=10),
                text_color="white",
                fg_color=COLORES["advertencia"],
                corner_radius=4,
                padx=8,
                pady=2
            )
            pass_badge.pack(side="left")
        
        # Derecha: Acciones
        acciones_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        acciones_frame.pack(side="right", padx=(10, 0))
        
        if usuario.get('activo') == 1:
            desactivar_btn = ctk.CTkButton(
                acciones_frame,
                text="Desactivar",
                command=lambda u=usuario: self.desactivar_usuario(u,),
                width=100,
                height=30,
                font=ctk.CTkFont(size=11),
                fg_color=COLORES["advertencia"],
                hover_color=COLORES["peligro"]
            )
            desactivar_btn.pack(pady=2)
        else:
            activar_btn = ctk.CTkButton(
                acciones_frame,
                text="Activar",
                command=lambda u=usuario: self.activar_usuario(u,),
                width=100,
                height=30,
                font=ctk.CTkFont(size=11),
                fg_color=COLORES["exito"],
                hover_color=COLORES["exito_hover"]
            )
            activar_btn.pack(pady=2)
    
    def abrir_crear_usuario(self):
        """Abre diálogo para crear usuario."""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Crear Nuevo Usuario")
        dialog.geometry("500x550")
        dialog.transient(self)
        dialog.grab_set()
        
        # Centrar
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (250)
        y = (dialog.winfo_screenheight() // 2) - (275)
        dialog.geometry(f'500x550+{x}+{y}')
        
        # Contenido
        main = ctk.CTkFrame(dialog, fg_color=COLORES["fondo_claro"])
        main.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        titulo = ctk.CTkLabel(
            main,
            text="➕ Crear Nuevo Usuario",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        titulo.pack(pady=(0, 20))
        
        # Formulario
        # Usuario
        ctk.CTkLabel(main, text="Usuario:", font=ctk.CTkFont(size=12, weight="bold"), anchor="w").pack(fill="x", pady=(5, 2))
        usuario_entry = ctk.CTkEntry(main, placeholder_text="Nombre de usuario", height=35,)
        usuario_entry.pack(fill="x", pady=(0, 10))
        
        # Nombre completo
        ctk.CTkLabel(main, text="Nombre Completo:", font=ctk.CTkFont(size=12, weight="bold"), anchor="w").pack(fill="x", pady=(5, 2))
        nombre_entry = ctk.CTkEntry(main, placeholder_text="Nombre y apellidos", height=35,)
        nombre_entry.pack(fill="x", pady=(0, 10))
        
        # Email
        ctk.CTkLabel(main, text="Email:", font=ctk.CTkFont(size=12, weight="bold"), anchor="w").pack(fill="x", pady=(5, 2))
        email_entry = ctk.CTkEntry(main, placeholder_text="correo@ejemplo.com", height=35,)
        email_entry.pack(fill="x", pady=(0, 10))
        
        # Configuración SMTP (opcional)
        smtp_frame = ctk.CTkFrame(main, fg_color=COLORES["fondo_claro"], border_width=1, border_color=COLORES["borde"])
        smtp_frame.pack(fill="x", pady=(10, 10))
        
        smtp_titulo = ctk.CTkLabel(
            smtp_frame,
            text="📧 Configuración de Email (Opcional)",
            font=ctk.CTkFont(size=11, weight="bold")
        )
        smtp_titulo.pack(pady=(10, 5))
        
        smtp_info = ctk.CTkLabel(
            smtp_frame,
            text="Configure para enviar contraseña por email",
            font=ctk.CTkFont(size=10),
            text_color=COLORES["texto_secundario"]
        )
        smtp_info.pack(pady=(0, 10))
        
        # Email remitente
        ctk.CTkLabel(smtp_frame, text="Email Remitente:", font=ctk.CTkFont(size=10),  anchor="w").pack(fill="x", padx=10, pady=(5, 2))
        smtp_email_entry = ctk.CTkEntry(smtp_frame, placeholder_text="tu_email@gmail.com", height=30,)
        smtp_email_entry.pack(fill="x", padx=10, pady=(0, 5))
        
        # Password del email
        ctk.CTkLabel(smtp_frame, text="Contraseña App de Gmail:", font=ctk.CTkFont(size=10), anchor="w").pack(fill="x", padx=10, pady=(5, 2))
        smtp_pass_entry = ctk.CTkEntry(smtp_frame, placeholder_text="Contraseña de aplicación", show="●", height=30,)
        smtp_pass_entry.pack(fill="x", padx=10, pady=(0, 10))
        
        # Botones
        botones_frame = ctk.CTkFrame(main, fg_color="transparent")
        botones_frame.pack(fill="x", pady=(20, 0))
        
        def crear():
            usuario = usuario_entry.get().strip()
            nombre = nombre_entry.get().strip()
            email = email_entry.get().strip()
            smtp_email = smtp_email_entry.get().strip()
            smtp_pass = smtp_pass_entry.get().strip()
            
            if not usuario or not nombre:
                messagebox.showerror("Error", "Usuario y nombre completo son obligatorios", parent=dialog)
                return
            
            # Configurar SMTP si se proporcionó
            smtp_config = None
            if smtp_email and smtp_pass:
                smtp_config = {
                    'servidor': 'smtp.gmail.com',
                    'puerto': 587,
                    'email': smtp_email,
                    'password': smtp_pass
                }
            
            # Crear usuario
            exito, password_temp, mensaje = self.db.agregar_usuario(
                usuario=usuario,
                nombre_completo=nombre,
                email=email if email else "",
                enviar_email=bool(smtp_config and email),
                smtp_config=smtp_config
            )
            
            if exito:
                messagebox.showinfo("Éxito", mensaje, parent=dialog)
                self.cargar_usuarios()
                dialog.destroy()
            else:
                messagebox.showerror("Error", mensaje, parent=dialog)
        
        crear_btn = ctk.CTkButton(
            botones_frame,
            text="Crear Usuario",
            command=crear,
            height=40,
            fg_color=COLORES["exito"],
            hover_color=COLORES["exito_hover"],
            font=ctk.CTkFont(size=13, weight="bold",)
        )
        crear_btn.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        cancelar_btn = ctk.CTkButton(
            botones_frame,
            text="Cancelar",
            command=dialog.destroy,
            height=40,
            fg_color=COLORES["borde"],
            hover_color=COLORES["texto_secundario"],)
        cancelar_btn.pack(side="right", fill="x", expand=True, padx=(5, 0))
    
    def activar_usuario(self, usuario):
        """Activa un usuario."""
        if self.db.activar_desactivar_usuario(usuario['usuario'], True):
            messagebox.showinfo("Éxito", f"Usuario '{usuario['usuario']}' activado", parent=self)
            self.cargar_usuarios()
    
    def desactivar_usuario(self, usuario):
        """Desactiva un usuario."""
        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Desactivar el usuario '{usuario['usuario']}'?\nNo podrá iniciar sesión.",
            parent=self
        )
        if respuesta:
            if self.db.activar_desactivar_usuario(usuario['usuario'], False):
                messagebox.showinfo("Éxito", f"Usuario '{usuario['usuario']}' desactivado", parent=self)
                self.cargar_usuarios()


class LoginWindow(ctk.CTkToplevel):
    """Ventana de inicio de sesión."""
    
    def __init__(self, parent, db):
        super().__init__(parent)
        
        self.db = db
        self.usuario_autenticado = None
        
        # Configuración de ventana
        self.title("ASEGURANZA - Inicio de Sesión")
        self.geometry("450x550")
        self.resizable(False, False)
        
        # Centrar ventana
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        # Hacer modal
        self.transient(parent)
        self.grab_set()
        
        # Crear interfaz
        self.crear_interfaz()
        
    def crear_interfaz(self):
        """Crea la interfaz de login."""
        # Frame principal
        main_frame = ctk.CTkFrame(self, fg_color=COLORES["fondo_claro"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Logo/Título
        titulo_frame = ctk.CTkFrame(main_frame, fg_color=COLORES["primario"],)
        titulo_frame.pack(fill="x", pady=(0, 30))
        
        titulo = ctk.CTkLabel(
            titulo_frame,
            text="🔐 ASEGURANZA",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="white"
        )
        titulo.pack(pady=20)
        
        subtitulo = ctk.CTkLabel(
            titulo_frame,
            text="Sistema de Gestión de Pólizas",
            font=ctk.CTkFont(size=14),
            text_color=COLORES["texto_secundario"]
        )
        subtitulo.pack(pady=(0, 15))
        
        # Frame de formulario
        form_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        form_frame.pack(fill="both", expand=True, padx=20)
        
        # Label de instrucción
        instruccion = ctk.CTkLabel(
            form_frame,
            text="Ingrese sus credenciales para continuar",
            font=ctk.CTkFont(size=12),
            text_color=COLORES["texto_secundario"]
        )
        instruccion.pack(pady=(0, 20))
        
        # Usuario
        usuario_label = ctk.CTkLabel(
            form_frame,
            text="Usuario:",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        usuario_label.pack(fill="x", pady=(10, 5))
        
        self.usuario_entry = ctk.CTkEntry(
            form_frame,
            placeholder_text="Ingrese su usuario",
            height=45,
            font=ctk.CTkFont(size=14,),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.usuario_entry.pack(fill="x", pady=(0, 15))
        self.usuario_entry.bind('<Return>', lambda e: self.password_entry.focus())
        
        # Contraseña
        password_label = ctk.CTkLabel(
            form_frame,
            text="Contraseña:",
            font=ctk.CTkFont(size=13, weight="bold"),
            anchor="w"
        )
        password_label.pack(fill="x", pady=(10, 5))
        
        self.password_entry = ctk.CTkEntry(
            form_frame,
            placeholder_text="Ingrese su contraseña",
            show="●",
            height=45,
            font=ctk.CTkFont(size=14,),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.password_entry.pack(fill="x", pady=(0, 25))
        self.password_entry.bind('<Return>', lambda e: self.iniciar_sesion())
        
        # Botón de inicio de sesión
        self.login_button = ctk.CTkButton(
            form_frame,
            text="Iniciar Sesión",
            command=self.iniciar_sesion,
            height=50,
            font=ctk.CTkFont(size=15, weight="bold",),
            fg_color=COLORES["primario"],
            hover_color=COLORES["primario_hover"],
            
        )
        self.login_button.pack(fill="x", pady=(0, 15))
        
        # Información de usuario por defecto
        info_frame = ctk.CTkFrame(form_frame, fg_color=COLORES["info"],)
        info_frame.pack(fill="x", pady=(20, 0))
        
        info_texto = ctk.CTkLabel(
            info_frame,
            text="ℹ️ Usuario por defecto:\nUsuario: admin | Contraseña: admin123",
            font=ctk.CTkFont(size=11),
            text_color="white",
            justify="center"
        )
        info_texto.pack(pady=10)
        
        # Focus inicial
        self.usuario_entry.focus()
    
    def iniciar_sesion(self):
        """Valida las credenciales e inicia sesión."""
        usuario = self.usuario_entry.get().strip()
        password = self.password_entry.get()
        
        if not usuario or not password:
            messagebox.showerror(
                "Error",
                "Por favor ingrese usuario y contraseña",
                parent=self
            )
            return
        
        # Verificar credenciales
        usuario_data = self.db.verificar_credenciales(usuario, password)
        
        if usuario_data:
            # Verificar si debe cambiar contraseña
            if usuario_data.get('cambiar_password') == 1:
                # Forzar cambio de contraseña
                cambio_window = CambiarPasswordWindow(
                    self,
                    self.db,
                    usuario_data,
                    obligatorio=True
                )
                self.wait_window(cambio_window)
                
                # Si cambió la contraseña, continuar
                if cambio_window.password_cambiado:
                    self.usuario_autenticado = usuario_data
                    self.destroy()
                else:
                    # Si no cambió la contraseña, no permitir continuar
                    messagebox.showwarning(
                        "Advertencia",
                        "Debe cambiar su contraseña para continuar",
                        parent=self
                    )
            else:
                self.usuario_autenticado = usuario_data
                self.destroy()
        else:
            messagebox.showerror(
                "Error de Autenticación",
                "Usuario o contraseña incorrectos",
                parent=self
            )
            self.password_entry.delete(0, 'end')
            self.password_entry.focus()



class AgendaTelefonicaWindow(ctk.CTkToplevel):
    """Ventana de Agenda Telefónica."""
    
    def __init__(self, parent, db):
        super().__init__(parent)
        self.db = db
        
        self.title("ASEGURANZA - Agenda Telefónica")
        self.geometry("900x700")
        self.minsize(800, 600)
        
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        self.crear_interfaz()
        self.cargar_contactos()
        
    def crear_interfaz(self):
        main_frame = ctk.CTkFrame(self, fg_color=COLORES["fondo_claro"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        header_frame = ctk.CTkFrame(main_frame, fg_color=COLORES["primario"])
        header_frame.pack(fill="x", pady=(0, 20))
        
        titulo = ctk.CTkLabel(
            header_frame,
            text="📒 AGENDA TELEFÓNICA",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        )
        titulo.pack(pady=15)
        
        self.tabview = ctk.CTkTabview(main_frame, fg_color="transparent")
        self.tabview.pack(fill="both", expand=True)
        
        self.tab_todos = self.tabview.add("Todos")
        self.tab_normales = self.tabview.add("Clientes Normales")
        self.tab_companias = self.tabview.add("Compañías")
        
        # Scrolables
        self.scroll_todos = ctk.CTkScrollableFrame(self.tab_todos, fg_color="transparent")
        self.scroll_todos.pack(fill="both", expand=True)
        
        self.scroll_normales = ctk.CTkScrollableFrame(self.tab_normales, fg_color="transparent")
        self.scroll_normales.pack(fill="both", expand=True)
        
        self.scroll_companias = ctk.CTkScrollableFrame(self.tab_companias, fg_color="transparent")
        self.scroll_companias.pack(fill="both", expand=True)
        
    def cargar_contactos(self):
        polizas = self.db.obtener_todas_polizas()
        # Group by phone to avoid duplicates if same client has multiple policies
        contactos = {}
        for p in polizas:
            tel = p.get('telefono', '').strip()
            if not tel:
                continue
            nombre = p.get('nombre_completo', '').strip()
            email = p.get('email', '').strip()
            tipo = p.get('tipo_cliente', 'Cliente Normal')
            
            key = (nombre.lower(), tel)
            if key not in contactos:
                contactos[key] = {
                    'nombre': nombre,
                    'telefono': tel,
                    'email': email,
                    'tipo': tipo
                }
                
        contactos_lista = sorted(list(contactos.values()), key=lambda x: x['nombre'].lower())
        
        for p in contactos_lista:
            self.crear_tarjeta(p, self.scroll_todos)
            if p['tipo'] == 'Compañía':
                self.crear_tarjeta(p, self.scroll_companias)
            else:
                self.crear_tarjeta(p, self.scroll_normales)
                
    def crear_tarjeta(self, contacto, parent_frame):
        card = ctk.CTkFrame(parent_frame, fg_color=COLORES["fondo"], border_width=1, border_color=COLORES["borde"])
        card.pack(fill="x", pady=5, padx=5)
        
        left_frame = ctk.CTkFrame(card, fg_color="transparent")
        left_frame.pack(side="left", fill="both", expand=True, padx=15, pady=10)
        
        nombre_lbl = ctk.CTkLabel(left_frame, text=f"👤 {contacto['nombre']}", font=ctk.CTkFont(size=14, weight="bold"))
        nombre_lbl.pack(anchor="w")
        
        if contacto['tipo'] == 'Compañía':
            badge = ctk.CTkLabel(left_frame, text="🏢 Compañía", font=ctk.CTkFont(size=10), text_color="white", fg_color=COLORES["info"], corner_radius=4, padx=6)
            badge.pack(anchor="w", pady=(2, 0))
        
        right_frame = ctk.CTkFrame(card, fg_color="transparent")
        right_frame.pack(side="right", padx=15, pady=10)
        
        ctk.CTkLabel(right_frame, text=f"📞 {contacto['telefono']}", font=ctk.CTkFont(size=13, weight="bold"), text_color=COLORES["exito"]).pack(anchor="e")
        if contacto['email']:
            ctk.CTkLabel(right_frame, text=f"✉️ {contacto['email']}", font=ctk.CTkFont(size=11), text_color=COLORES["texto_secundario"]).pack(anchor="e")



class SeguroApp(ctk.CTk):
    """Aplicación principal de gestión de pólizas."""
    
    def __init__(self, usuario_autenticado=None, db_path="seguros.db"):
        super().__init__()
        
        # Usuario autenticado
        self.usuario_autenticado = usuario_autenticado
        
        # Configuración de la ventana
        self.title("ASEGURANZA - Sistema de Gestión de Pólizas")
        self.geometry("1450x850")
        self.minsize(1200, 700)
        
        # Centrar ventana
        self.center_window()
        
        # Tema oscuro profesional
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        # Base de datos
        self.db = DatabaseManager(db_name=db_path)
        
        # Variables
        self.poliza_seleccionada_id = None
        self.modo_edicion = False
        
        # Configurar grid principal
        self.grid_columnconfigure(0, weight=0)  # Header
        self.grid_columnconfigure(1, weight=1)  # Contenido
        self.grid_rowconfigure(0, weight=0)     # Header
        self.grid_rowconfigure(1, weight=1)     # Contenido
        self.grid_rowconfigure(2, weight=0)     # Footer
        
        # Crear interfaz
        self.crear_header()
        self.crear_panel_izquierdo()
        self.crear_panel_derecho()
        self.crear_footer()
        
        # Cargar pólizas iniciales
        self.cargar_polizas()
        
        # Mostrar alertas de vencimiento con un pequeño retraso
        self.after(500, self.mostrar_alertas_vencimiento)
        
        # Iniciar auto-refresco para la red local
        self.ultima_act_db = ""
        self.after(3000, self.iniciar_auto_refresh)
        
    def iniciar_auto_refresh(self):
        """Verifica cambios en la BD y actualiza la lista si alguien más modifica algo en la red local."""
        if not self.search_var.get():
            try:
                nueva_act = self.db.obtener_ultima_actualizacion()
                if self.ultima_act_db and nueva_act != self.ultima_act_db:
                    self.cargar_polizas()
                self.ultima_act_db = nueva_act
            except Exception:
                pass
        self.after(8000, self.iniciar_auto_refresh)
    
    def mostrar_alertas_vencimiento(self):
        """Muestra un popup con las pólizas que están por vencer (15, 7 o 2 días)."""
        alertas = self.db.obtener_alertas_vencimientos()
        if not alertas:
            return
            
        mensaje = "Las siguientes pólizas están próximas a vencer:\n\n"
        for alerta in alertas:
            nombre = alerta.get('nombre_completo', 'Sin nombre')
            pol = alerta.get('numero_poliza') or 'Sin póliza'
            
            obs = []
            if alerta.get('alerta_cobertura') is not None:
                dias = alerta['alerta_cobertura']
                if dias == 0: obs.append("Cobertura termina HOY")
                else: obs.append(f"Cobertura en {abs(dias)} días")
                
            if alerta.get('alerta_pago') is not None:
                dias = alerta['alerta_pago']
                if dias == 0: obs.append("Pago vence HOY")
                else: obs.append(f"Pago en {abs(dias)} días")
                
            mensaje += f"• {nombre} ({pol}): {', '.join(obs)}\n"
            
        messagebox.showwarning(
            "Alertas de Vencimiento",
            mensaje,
            parent=self
        )
    
    def center_window(self):
        """Centra la ventana en la pantalla."""
        self.update_idletasks()
        width = 1450
        height = 850
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
    
    def crear_header(self):
        """Crea el header profesional de la aplicación."""
        header_frame = ctk.CTkFrame(
            self,
            height=90,
            fg_color=COLORES["primario"],
            corner_radius=0
        )
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=0, pady=0)
        header_frame.grid_propagate(False)
        
        # Logo y título
        titulo_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        titulo_frame.pack(side="left", padx=40, pady=20)
        
        # Icono de la aplicación (simulado con emoji)
        icono = ctk.CTkLabel(
            titulo_frame,
            text="🚗",
            font=ctk.CTkFont(size=40)
        )
        icono.pack(side="left", padx=(0, 15))
        
        # Título y subtítulo
        texto_frame = ctk.CTkFrame(titulo_frame, fg_color="transparent")
        texto_frame.pack(side="left")
        
        titulo = ctk.CTkLabel(
            texto_frame,
            text="ASEGURANZA",
            font=ctk.CTkFont(size=28, weight="bold"),
            text_color="white"
        )
        titulo.pack(anchor="w")
        
        subtitulo = ctk.CTkLabel(
            texto_frame,
            text="Sistema de Gestión de Pólizas de Seguros",
            font=ctk.CTkFont(size=13),
            text_color=COLORES["texto_secundario"]
        )
        subtitulo.pack(anchor="w")
        
        # Información del sistema (derecha)
        info_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        info_frame.pack(side="right", padx=40, pady=20)
        
        # Usuario logueado
        if self.usuario_autenticado:
            usuario_label = ctk.CTkLabel(
                info_frame,
                text=f"👤 Usuario: {self.usuario_autenticado.get('nombre_completo', self.usuario_autenticado.get('usuario', 'N/A'))}",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="white"
            )
            usuario_label.pack(anchor="e", pady=(0, 5))
            
            # Botones de acción
            botones_frame = ctk.CTkFrame(info_frame, fg_color="transparent")
            botones_frame.pack(anchor="e")
            
            cambiar_pass_btn = ctk.CTkButton(
                botones_frame,
                text="🔑 Cambiar Contraseña",
                command=self.abrir_cambiar_password,
                width=150,
                height=28,
                font=ctk.CTkFont(size=11,),
                fg_color=COLORES["info"],
                hover_color=COLORES["primario_hover"]
            )
            cambiar_pass_btn.pack(side="left", padx=(0, 5))
            
            # Solo el admin puede gestionar usuarios
            if self.usuario_autenticado.get('usuario') == 'admin':
                gestion_usuarios_btn = ctk.CTkButton(
                    botones_frame,
                    text="👥 Gestión de Usuarios",
                    command=self.abrir_gestion_usuarios,
                    width=150,
                    height=28,
                    font=ctk.CTkFont(size=11,),
                    fg_color=COLORES["exito"],
                    hover_color=COLORES["exito_hover"]
                )
                gestion_usuarios_btn.pack(side="left", padx=(0, 5))

            agenda_btn = ctk.CTkButton(
                botones_frame,
                text="📒 Agenda Telefónica",
                command=self.abrir_agenda,
                width=150,
                height=28,
                font=ctk.CTkFont(size=11),
                fg_color=COLORES["advertencia"],
                hover_color="#d97706"
            )
            agenda_btn.pack(side="left")

            # Botón Salir/Cerrar Sesión
            salir_btn = ctk.CTkButton(
                botones_frame,
                text="🚪 Cerrar Sesión",
                command=self.salir_programa,
                width=120,
                height=28,
                font=ctk.CTkFont(size=11),
                fg_color=COLORES["peligro"],
                hover_color=COLORES["peligro_hover"]
            )
            salir_btn.pack(side="left", padx=(5, 0))
        
        # Fecha actual
        fecha_actual = datetime.now().strftime("%d/%m/%Y")
        fecha_label = ctk.CTkLabel(
            info_frame,
            text=f"📅 {fecha_actual}",
            font=ctk.CTkFont(size=12),
            text_color=COLORES["texto_secundario"]
        )
        fecha_label.pack(anchor="e", pady=(3, 0))
        
        # Total de pólizas
        self.total_polizas_label = ctk.CTkLabel(
            info_frame,
            text="📊 Total de pólizas: 0",
            font=ctk.CTkFont(size=12),
            text_color=COLORES["texto_secundario"]
        )
        self.total_polizas_label.pack(anchor="e", pady=(3, 0))
    
    def crear_footer(self):
        """Crea el footer de la aplicación."""
        footer_frame = ctk.CTkFrame(
            self,
            height=35,
            fg_color=COLORES["fondo_claro"],
            corner_radius=0
        )
        footer_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=0, pady=0)
        footer_frame.grid_propagate(False)
        
        # Versión del sistema
        version_label = ctk.CTkLabel(
            footer_frame,
            text="Versión 1.0.0 | © 2026 ASEGURANZA | Todos los derechos reservados",
            font=ctk.CTkFont(size=11),
            text_color=COLORES["texto_secundario"]
        )
        version_label.pack(side="left", padx=20)
        
        # Estado del sistema
        self.estado_label = ctk.CTkLabel(
            footer_frame,
            text="✓ Sistema operativo",
            font=ctk.CTkFont(size=11),
            text_color=COLORES["exito"]
        )
        self.estado_label.pack(side="right", padx=20)
    
    def crear_panel_izquierdo(self):
        """Crea el panel izquierdo con búsqueda y lista de resultados."""
        # Frame principal izquierdo con borde visual
        self.panel_izquierdo = ctk.CTkFrame(
            self,
            width=440,
            fg_color=COLORES["fondo_claro"],
            border_width=1,
            border_color=COLORES["borde"]
        )
        self.panel_izquierdo.grid(row=1, column=0, sticky="nsew", padx=(25, 12), pady=(25, 12))
        self.panel_izquierdo.grid_propagate(False)
        
        # Header del panel
        panel_header = ctk.CTkFrame(
            self.panel_izquierdo,
            fg_color=COLORES["primario"],

            height=60
        )
        panel_header.pack(pady=(15, 10), padx=15, fill="x")
        panel_header.pack_propagate(False)
        
        titulo = ctk.CTkLabel(
            panel_header,
            text="📋 REGISTRO DE PÓLIZAS",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="white"
        )
        titulo.pack(pady=15)
        
        # Barra de búsqueda mejorada
        search_frame = ctk.CTkFrame(self.panel_izquierdo, fg_color="transparent")
        search_frame.pack(pady=(10, 5), padx=15, fill="x")
        
        search_label = ctk.CTkLabel(
            search_frame,
            text="Búsqueda Universal:",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLORES["texto"]
        )
        search_label.pack(anchor="w", pady=(0, 5))
        
        self.search_var = ctk.StringVar()
        self.search_var.trace_add('write', self.buscar_en_tiempo_real)
        
        search_entry = ctk.CTkEntry(
            search_frame,
            placeholder_text="🔍 Buscar por cualquier campo...",
            textvariable=self.search_var,
            height=45,
            font=ctk.CTkFont(size=13,),
            border_width=2,
            border_color=COLORES["borde"]
        )
        search_entry.pack(fill="x")
        
        # Contador de resultados
        self.contador_label = ctk.CTkLabel(
            self.panel_izquierdo,
            text="📊 Mostrando 0 pólizas",
            font=ctk.CTkFont(size=11),
            text_color=COLORES["texto_secundario"]
        )
        self.contador_label.pack(anchor="w", padx=15, pady=(5, 10))
        
        # Botón Nueva Póliza con diseño mejorado
        btn_nueva = ctk.CTkButton(
            self.panel_izquierdo,
            text="➕ NUEVA PÓLIZA",
            command=self.nueva_poliza,
            height=50,
            font=ctk.CTkFont(size=15, weight="bold",),
            fg_color=COLORES["exito"],
            hover_color=COLORES["exito_hover"],

            border_width=0
        )
        btn_nueva.pack(pady=(5, 15), padx=15, fill="x")
        
        # Separador visual
        separador = ctk.CTkFrame(
            self.panel_izquierdo,
            height=2,
            fg_color=COLORES["borde"]
        )
        separador.pack(fill="x", padx=15, pady=(0, 10))
        
        # Frame scrollable para lista de pólizas
        self.lista_frame = ctk.CTkScrollableFrame(
            self.panel_izquierdo,
            fg_color="transparent"
        )
        self.lista_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
    
    def crear_panel_derecho(self):
        """Crea el panel derecho con detalles y formulario."""
        # Frame principal derecho con borde
        self.panel_derecho = ctk.CTkFrame(
            self,
            fg_color=COLORES["fondo_claro"],
            border_width=1,
            border_color=COLORES["borde"]
        )
        self.panel_derecho.grid(row=1, column=1, sticky="nsew", padx=(12, 25), pady=(25, 12))
        
        # Header del formulario
        form_header = ctk.CTkFrame(
            self.panel_derecho,
            fg_color=COLORES["secundario"],

            height=70
        )
        form_header.pack(pady=(15, 10), padx=15, fill="x")
        form_header.pack_propagate(False)
        
        # Frame para título y estado
        header_content = ctk.CTkFrame(form_header, fg_color="transparent")
        header_content.pack(fill="both", expand=True, padx=20, pady=10)
        
        self.form_titulo = ctk.CTkLabel(
            header_content,
            text="📄 DETALLES DE LA PÓLIZA",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="white"
        )
        self.form_titulo.pack(anchor="w")
        
        self.form_subtitulo = ctk.CTkLabel(
            header_content,
            text="Complete los campos para registrar una nueva póliza",
            font=ctk.CTkFont(size=12),
            text_color=COLORES["texto_secundario"]
        )
        self.form_subtitulo.pack(anchor="w", pady=(2, 0))
        
        # Frame scrollable para el formulario
        self.form_frame = ctk.CTkScrollableFrame(
            self.panel_derecho,
            fg_color="transparent"
        )
        self.form_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # BLOQUE 1: Información Personal
        self.crear_seccion("👤 INFORMACIÓN PERSONAL", self.form_frame, COLORES["info"])
        
        # Crear campos en 3 columnas para mejor uso del espacio
        row1 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row1.pack(fill="x", pady=5)
        
        col1 = ctk.CTkFrame(row1, fg_color="transparent")
        col1.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col_tipo = ctk.CTkFrame(row1, fg_color="transparent")
        col_tipo.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col2 = ctk.CTkFrame(row1, fg_color="transparent")
        col2.pack(side="left", fill="both", expand=True)
        
        self.nombre_entry = self.crear_campo("Nombre Completo *", col1, True)
        
        self.crear_label("Tipo de Cliente", col_tipo)
        self.tipo_cliente_combo = ctk.CTkComboBox(
            col_tipo,
            values=["Cliente Normal", "Compañía"],
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.tipo_cliente_combo.pack(fill="x", pady=(0, 10))
        self.tipo_cliente_combo.set("Cliente Normal")
        
        self.rfc_entry = self.crear_campo("RFC (10 caracteres)", col2, True)
        
        # Fila 2
        row2 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row2.pack(fill="x", pady=5)
        
        col3 = ctk.CTkFrame(row2, fg_color="transparent")
        col3.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col_tel = ctk.CTkFrame(row2, fg_color="transparent")
        col_tel.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col4 = ctk.CTkFrame(row2, fg_color="transparent")
        col4.pack(side="left", fill="both", expand=True)
        
        self.email_entry = self.crear_campo("Email", col3, True)
        self.telefono_entry = self.crear_campo("Teléfono", col_tel, True)
        self.calle_entry = self.crear_campo("Calle", col4, True)
        
        # Fila 3
        row3 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row3.pack(fill="x", pady=5)
        
        col5 = ctk.CTkFrame(row3, fg_color="transparent")
        col5.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col6 = ctk.CTkFrame(row3, fg_color="transparent")
        col6.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col7 = ctk.CTkFrame(row3, fg_color="transparent")
        col7.pack(side="left", fill="both", expand=True)
        
        self.numero_entry = self.crear_campo("Número", col5, True)
        
        # Código Postal con validación en tiempo real para API
        self.crear_label("Código Postal", col6)
        self.cp_var = ctk.StringVar()
        self.cp_var.trace_add('write', self.validar_cp_y_buscar)
        self.cp_entry = ctk.CTkEntry(
            col6,
            height=40,
            font=ctk.CTkFont(size=13),
            corner_radius=6,
            border_width=2,
            border_color=COLORES["borde"],
            textvariable=self.cp_var
        )
        self.cp_entry.pack(fill="x", pady=(0, 10))
        
        # Colonia ahora es un ComboBox para seleccionar entre las devueltas por el CP
        self.crear_label("Colonia", col7)
        self.colonia_combo = ctk.CTkComboBox(
            col7,
            values=[],
            height=40,
            font=ctk.CTkFont(size=13),
            corner_radius=6,
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.colonia_combo.pack(fill="x", pady=(0, 10))
        self.colonia_combo.set("")
        
        # Fila 4 - Estado y Municipio con ComboBox
        row4 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row4.pack(fill="x", pady=5)
        
        col8 = ctk.CTkFrame(row4, fg_color="transparent")
        col8.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col9 = ctk.CTkFrame(row4, fg_color="transparent")
        col9.pack(side="left", fill="both", expand=True)
        
        # Estado con ComboBox (primero - izquierda)
        self.crear_label("Estado", col8)
        self.estado_combo = ctk.CTkComboBox(
            col8,
            values=ESTADOS_MEXICO,
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"],
            command=self.actualizar_municipios
        )
        self.estado_combo.pack(fill="x", pady=(0, 10))
        self.estado_combo.set("")  # Vacío por defecto
        
        # Municipio con ComboBox (después - derecha)
        self.crear_label("Municipio", col9)
        self.municipio_combo = ctk.CTkComboBox(
            col9,
            values=[],  # Se llena cuando se selecciona un estado
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.municipio_combo.pack(fill="x", pady=(0, 10))
        self.municipio_combo.set("")  # Vacío por defecto
        
        # BLOQUE 2: Datos de Póliza
        self.crear_seccion("📋 DATOS DE PÓLIZA", self.form_frame, COLORES["secundario"])
        
        # Fila de póliza
        row5 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row5.pack(fill="x", pady=5)
        
        col10 = ctk.CTkFrame(row5, fg_color="transparent")
        col10.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col11 = ctk.CTkFrame(row5, fg_color="transparent")
        col11.pack(side="left", fill="both", expand=True)
        
        self.compania_entry = self.crear_campo("Compañía Aseguradora", col10, True)
        self.numero_poliza_entry = self.crear_campo("Número de Póliza", col11, True)
        
        # Fila de endoso
        row6 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row6.pack(fill="x", pady=5)
        
        col12 = ctk.CTkFrame(row6, fg_color="transparent")
        col12.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col13 = ctk.CTkFrame(row6, fg_color="transparent")
        col13.pack(side="left", fill="both", expand=True)
        
        self.endoso_entry = self.crear_campo("Endoso", col12, True)
        self.inciso_entry = self.crear_campo("Inciso", col13, True)
        
        # Fila de vigencia
        row7 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row7.pack(fill="x", pady=5)
        
        col14 = ctk.CTkFrame(row7, fg_color="transparent")
        col14.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col15 = ctk.CTkFrame(row7, fg_color="transparent")
        col15.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col16 = ctk.CTkFrame(row7, fg_color="transparent")
        col16.pack(side="left", fill="both", expand=True)
        
        self.vigencia_inicio_entry = self.crear_campo("Vigencia Inicio (DD-MM-AAAA)", col14, True)
        self.vigencia_fin_entry = self.crear_campo("Vigencia Fin (DD-MM-AAAA)", col15, True)
        
        # Campo de vencimiento con label de días restantes
        self.crear_label("Vencimiento Pago (DD-MM-AAAA)", col16)
        self.fecha_vencimiento_entry = ctk.CTkEntry(
            col16,
            height=40,
            font=ctk.CTkFont(size=13,),
            border_width=2,
            border_color=COLORES["borde"],
            
        )
        self.fecha_vencimiento_entry.pack(fill="x", pady=(0, 5))
        self.fecha_vencimiento_entry.bind('<FocusOut>', self.calcular_dias_vencimiento)
        self.fecha_vencimiento_entry.bind('<KeyRelease>', self.calcular_dias_vencimiento)
        self.vigencia_fin_entry.bind('<FocusOut>', self.calcular_dias_vencimiento)
        self.vigencia_fin_entry.bind('<KeyRelease>', self.calcular_dias_vencimiento)
        
        # Label para mostrar días restantes
        self.dias_restantes_label = ctk.CTkLabel(
            col16,
            text="",
            font=ctk.CTkFont(size=10),
            text_color=COLORES["texto_secundario"],
            justify="left"
        )
        self.dias_restantes_label.pack(anchor="w", pady=(0, 10))
        
        # Forma de pago, tipo de moneda y prima
        row8 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row8.pack(fill="x", pady=5)
        
        col17 = ctk.CTkFrame(row8, fg_color="transparent")
        col17.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col18 = ctk.CTkFrame(row8, fg_color="transparent")
        col18.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col19 = ctk.CTkFrame(row8, fg_color="transparent")
        col19.pack(side="left", fill="both", expand=True)
        
        self.crear_label("Forma de Pago", col17)
        self.forma_pago_combo = ctk.CTkComboBox(
            col17,
            values=FORMAS_PAGO,
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.forma_pago_combo.pack(fill="x", pady=(0, 10))
        
        self.crear_label("Tipo de Moneda", col18)
        self.tipo_moneda_combo = ctk.CTkComboBox(
            col18,
            values=TIPOS_MONEDA,
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.tipo_moneda_combo.pack(fill="x", pady=(0, 10))
        self.tipo_moneda_combo.set("MXN")
        
        self.prima_entry = self.crear_campo("Prima Total ($)", col19, True)
        
        # BLOQUE 3: Datos del Vehículo
        self.crear_seccion("🚗 DATOS DEL VEHÍCULO", self.form_frame, COLORES["advertencia"])
        
        # Fila vehículo 1
        row9 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row9.pack(fill="x", pady=5)
        
        col20 = ctk.CTkFrame(row9, fg_color="transparent")
        col20.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col21 = ctk.CTkFrame(row9, fg_color="transparent")
        col21.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col22 = ctk.CTkFrame(row9, fg_color="transparent")
        col22.pack(side="left", fill="both", expand=True)
        
        self.marca_entry = self.crear_campo("Marca", col20, True)
        self.modelo_entry = self.crear_campo("Modelo", col21, True)
        self.anio_entry = self.crear_campo("Año", col22, True)
        
        # Fila vehículo 2
        row10 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row10.pack(fill="x", pady=5)
        
        col23 = ctk.CTkFrame(row10, fg_color="transparent")
        col23.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col24 = ctk.CTkFrame(row10, fg_color="transparent")
        col24.pack(side="left", fill="both", expand=True)
        
        self.tipo_version_entry = self.crear_campo("Tipo/Versión", col23, True)
        self.placas_entry = self.crear_campo("Placas", col24, True)
        
        # Fila vehículo 3 - Uso, Servicio
        row11 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row11.pack(fill="x", pady=5)
        
        col25 = ctk.CTkFrame(row11, fg_color="transparent")
        col25.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col26 = ctk.CTkFrame(row11, fg_color="transparent")
        col26.pack(side="left", fill="both", expand=True)
        
        self.crear_label("Uso", col25)
        self.uso_combo = ctk.CTkComboBox(
            col25,
            values=USOS_VEHICULO,
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.uso_combo.pack(fill="x", pady=(0, 10))
        
        self.crear_label("Servicio", col26)
        self.servicio_combo = ctk.CTkComboBox(
            col26,
            values=SERVICIOS_VEHICULO,
            height=40,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"]
        )
        self.servicio_combo.pack(fill="x", pady=(0, 10))
        
        # Fila vehículo 4 - Serie/VIN, Motor, Último Movimiento
        row12 = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        row12.pack(fill="x", pady=5)
        
        col27 = ctk.CTkFrame(row12, fg_color="transparent")
        col27.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col28 = ctk.CTkFrame(row12, fg_color="transparent")
        col28.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        col29 = ctk.CTkFrame(row12, fg_color="transparent")
        col29.pack(side="left", fill="both", expand=True)
        
        self.serie_vin_entry = self.crear_campo("Serie/VIN", col27, True)
        self.motor_entry = self.crear_campo("Motor", col28, True)
        self.ultimo_movimiento_entry = self.crear_campo("Último Movimiento", col29, True)
        
        # BLOQUE 4: Coberturas Contratadas
        self.crear_seccion("📊 COBERTURAS CONTRATADAS", self.form_frame, COLORES["info"])
        
        # Frame para la tabla de coberturas
        self.coberturas_frame = ctk.CTkFrame(
            self.form_frame,
            fg_color=COLORES["fondo_claro"],

            border_width=1,
            border_color=COLORES["borde"]
        )
        self.coberturas_frame.pack(fill="x", pady=(0, 15))
        
        # Headers de la tabla
        headers_row = ctk.CTkFrame(self.coberturas_frame, fg_color="transparent")
        headers_row.pack(fill="x", padx=10, pady=(10, 5))
        
        ctk.CTkLabel(
            headers_row,
            text="Cobertura",
            font=ctk.CTkFont(size=11, weight="bold"),
            width=200,
            anchor="w"
        ).pack(side="left", padx=5)
        
        ctk.CTkLabel(
            headers_row,
            text="Suma Asegurada",
            font=ctk.CTkFont(size=11, weight="bold"),
            width=150,
            anchor="w"
        ).pack(side="left", padx=5)
        
        ctk.CTkLabel(
            headers_row,
            text="Deducible",
            font=ctk.CTkFont(size=11, weight="bold"),
            width=120,
            anchor="w"
        ).pack(side="left", padx=5)
        
        ctk.CTkLabel(
            headers_row,
            text="Prima",
            font=ctk.CTkFont(size=11, weight="bold"),
            width=100,
            anchor="w"
        ).pack(side="left", padx=5)
        
        # Frame scrollable para las coberturas
        self.coberturas_lista_frame = ctk.CTkScrollableFrame(
            self.coberturas_frame,
            fg_color="transparent",
            height=150
        )
        self.coberturas_lista_frame.pack(fill="both", expand=True, padx=10, pady=(0, 5))
        
        # Botón para agregar cobertura
        btn_agregar_cobertura = ctk.CTkButton(
            self.coberturas_frame,
            text="➕ Agregar Cobertura",
            command=self.agregar_fila_cobertura,
            height=35,
            font=ctk.CTkFont(size=12,),
            fg_color=COLORES["exito"],
            hover_color=COLORES["exito_hover"]
        )
        btn_agregar_cobertura.pack(pady=10)
        
        # BLOQUE 5: Comentarios
        self.crear_seccion("💬 COMENTARIOS", self.form_frame, COLORES["secundario"])
        
        self.crear_label("Observaciones Adicionales", self.form_frame)
        self.comentarios_text = ctk.CTkTextbox(
            self.form_frame,
            height=100,
            font=ctk.CTkFont(size=13),
            border_width=2,
            border_color=COLORES["borde"],
            
        )
        self.comentarios_text.pack(fill="x", pady=(0, 15))
        
        # Botones de acción
        self.crear_botones_accion()
    
    def crear_seccion(self, titulo, parent, color=None):
        """Crea un título de sección con diseño mejorado."""
        if color is None:
            color = COLORES["secundario"]
            
        # Frame de la sección con fondo de color
        seccion_frame = ctk.CTkFrame(
            parent,
            fg_color=color,

            height=45
        )
        seccion_frame.pack(fill="x", pady=(20, 15))
        seccion_frame.pack_propagate(False)
        
        label = ctk.CTkLabel(
            seccion_frame,
            text=titulo,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="white"
        )
        label.pack(side="left", padx=15, pady=10)
    
    def crear_label(self, texto, parent):
        """Crea una etiqueta mejorada."""
        label = ctk.CTkLabel(
            parent,
            text=texto,
            font=ctk.CTkFont(size=11, weight="bold"),
            anchor="w",
            text_color=COLORES["texto"]
        )
        label.pack(anchor="w", pady=(8, 5))
    
    def crear_campo(self, placeholder, parent, mejorado=False):
        """Crea un campo de entrada mejorado."""
        self.crear_label(placeholder, parent)
        entry = ctk.CTkEntry(
            parent,
            height=40,
            font=ctk.CTkFont(size=13,),
            border_width=2,
            border_color=COLORES["borde"],
            
        )
        entry.pack(fill="x", pady=(0, 10))
        return entry
    
    def crear_botones_accion(self):
        """Crea los botones de acción mejorados."""
        # Separador visual
        separador = ctk.CTkFrame(
            self.form_frame,
            height=2,
            fg_color=COLORES["borde"]
        )
        separador.pack(fill="x", pady=(20, 20))
        
        # Título de acciones
        titulo_acciones = ctk.CTkLabel(
            self.form_frame,
            text="⚙️ ACCIONES",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLORES["texto"]
        )
        titulo_acciones.pack(anchor="w", pady=(0, 15))
        
        btn_frame = ctk.CTkFrame(self.form_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 20))
        
        # Primera fila de botones
        btn_row1 = ctk.CTkFrame(btn_frame, fg_color="transparent")
        btn_row1.pack(fill="x", pady=(0, 10))
        
        # Botón Guardar
        self.btn_guardar = ctk.CTkButton(
            btn_row1,
            text="💾 GUARDAR PÓLIZA",
            command=self.guardar_poliza,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold",),
            fg_color=COLORES["exito"],
            hover_color=COLORES["exito_hover"],

            border_width=0
        )
        self.btn_guardar.pack(side="left", expand=True, fill="x", padx=(0, 8))
        
        # Botón Cancelar
        self.btn_cancelar = ctk.CTkButton(
            btn_row1,
            text="❌ CANCELAR",
            command=self.cancelar_edicion,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold",),
            fg_color="#64748b",
            hover_color="#475569",

            border_width=0
        )
        self.btn_cancelar.pack(side="left", expand=True, fill="x")
        
        # Segunda fila de botones
        btn_row2 = ctk.CTkFrame(btn_frame, fg_color="transparent")
        btn_row2.pack(fill="x")
        
        # Botón Exportar
        self.btn_exportar = ctk.CTkButton(
            btn_row2,
            text="📊 EXPORTAR A EXCEL",
            command=self.exportar_excel,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold",),
            fg_color=COLORES["secundario"],
            hover_color=COLORES["primario_hover"],

            border_width=0
        )
        self.btn_exportar.pack(side="left", expand=True, fill="x", padx=(0, 8))
        
        # Botón Eliminar
        self.btn_eliminar = ctk.CTkButton(
            btn_row2,
            text="🗑️ ELIMINAR PÓLIZA",
            command=self.eliminar_poliza,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold",),
            fg_color=COLORES["peligro"],
            hover_color=COLORES["peligro_hover"],

            border_width=0
        )
        self.btn_eliminar.pack(side="left", expand=True, fill="x")
    
    def actualizar_municipios(self, estado_seleccionado):
        """Actualiza la lista de municipios según el estado seleccionado."""
        if estado_seleccionado in MUNICIPIOS_POR_ESTADO:
            municipios = MUNICIPIOS_POR_ESTADO[estado_seleccionado]
            self.municipio_combo.configure(values=municipios)
            self.municipio_combo.set("")  # Limpiar selección
        else:
            self.municipio_combo.configure(values=[])
            self.municipio_combo.set("")
    
    def calcular_dias_vencimiento(self, event=None):
        """Calcula los días restantes de cobertura y hasta el vencimiento del pago."""
        fecha_fin_str = self.vigencia_fin_entry.get().strip()
        fecha_pago_str = self.fecha_vencimiento_entry.get().strip()
        
        if not fecha_fin_str and not fecha_pago_str:
            self.dias_restantes_label.configure(text="")
            return
            
        texto_final = ""
        color_final = COLORES["texto_secundario"]
        fecha_hoy = datetime.now()
        error = False
        
        try:
            # Calcular Cobertura (Vigencia Fin)
            if fecha_fin_str:
                fecha_fin = datetime.strptime(fecha_fin_str, "%d-%m-%Y")
                dias_cobertura = (fecha_fin - fecha_hoy).days
                
                if dias_cobertura < 0:
                    texto_final += f"⚠️ Cobertura TERMINADA (hace {abs(dias_cobertura)} días)\n"
                    color_final = COLORES["peligro"]
                elif dias_cobertura == 0:
                    texto_final += "⚠️ La cobertura TERMINA HOY\n"
                    color_final = COLORES["peligro"]
                elif dias_cobertura <= 15:
                    texto_final += f"⚠️ Restan {dias_cobertura} días de cobertura\n"
                    color_final = COLORES["advertencia"]
                else:
                    texto_final += f"✓ Restan {dias_cobertura} días de cobertura\n"
                    color_final = COLORES["exito"]
                    
            # Calcular Vencimiento de Pago
            if fecha_pago_str:
                fecha_pago = datetime.strptime(fecha_pago_str, "%d-%m-%Y")
                dias_pago = (fecha_pago - fecha_hoy).days
                
                if dias_pago < 0:
                    texto_final += f"⚠️ Pago VENCIDO (hace {abs(dias_pago)} días)"
                    color_final = COLORES["peligro"]
                elif dias_pago == 0:
                    texto_final += "⚠️ El pago VENCE HOY"
                    color_final = COLORES["peligro"]
                elif dias_pago <= 15:
                    texto_final += f"⚠️ Faltan {dias_pago} días para el pago"
                    if color_final != COLORES["peligro"]:
                        color_final = COLORES["advertencia"]
                else:
                    texto_final += f"✓ Faltan {dias_pago} días para el pago"
                    
        except ValueError:
            texto_final = "❌ Formato inválido (DD-MM-AAAA)"
            color_final = COLORES["peligro"]
            error = True
            
        if not texto_final and not error:
            texto_final = "Ingrese fechas (DD-MM-AAAA)"
            
        self.dias_restantes_label.configure(text=texto_final.strip(), text_color=color_final)
    
    def agregar_fila_cobertura(self):
        """Agrega una nueva fila para ingresar una cobertura."""
        fila = ctk.CTkFrame(self.coberturas_lista_frame, fg_color="transparent")
        fila.pack(fill="x", pady=2)
        
        # Campo de cobertura
        cobertura_entry = ctk.CTkEntry(
            fila,
            width=200,
            height=35,
            font=ctk.CTkFont(size=11,),
            placeholder_text="Ej: R.C. Daños a Terceros"
        )
        cobertura_entry.pack(side="left", padx=5)
        
        # Campo de suma asegurada
        suma_entry = ctk.CTkEntry(
            fila,
            width=150,
            height=35,
            font=ctk.CTkFont(size=11,),
            placeholder_text="$0.00"
        )
        suma_entry.pack(side="left", padx=5)
        
        # Campo de deducible
        deducible_entry = ctk.CTkEntry(
            fila,
            width=120,
            height=35,
            font=ctk.CTkFont(size=11,),
            placeholder_text="5%"
        )
        deducible_entry.pack(side="left", padx=5)
        
        # Campo de prima
        prima_entry = ctk.CTkEntry(
            fila,
            width=100,
            height=35,
            font=ctk.CTkFont(size=11,),
            placeholder_text="$0.00"
        )
        prima_entry.pack(side="left", padx=5)
        
        # Botón para eliminar fila
        btn_eliminar = ctk.CTkButton(
            fila,
            text="🗑️",
            width=40,
            height=35,
            fg_color=COLORES["peligro"],
            hover_color=COLORES["peligro_hover"],
            command=lambda: fila.destroy()
        )
        btn_eliminar.pack(side="left", padx=5)
    
    def obtener_coberturas(self):
        """Obtiene todas las coberturas ingresadas y las convierte a JSON."""
        coberturas = []
        
        for fila in self.coberturas_lista_frame.winfo_children():
            widgets = [w for w in fila.winfo_children() if isinstance(w, ctk.CTkEntry)]
            
            if len(widgets) >= 4:
                cobertura = widgets[0].get()
                suma = widgets[1].get()
                deducible = widgets[2].get()
                prima = widgets[3].get()
                
                if cobertura:  # Solo agregar si tiene nombre de cobertura
                    coberturas.append({
                        "cobertura": cobertura,
                        "suma_asegurada": suma,
                        "deducible": deducible,
                        "prima": prima
                    })
        
        return json.dumps(coberturas, ensure_ascii=False) if coberturas else ""
    
    def cargar_coberturas(self, coberturas_json):
        """Carga las coberturas desde JSON y las muestra en la tabla."""
        # Limpiar coberturas existentes
        for widget in self.coberturas_lista_frame.winfo_children():
            widget.destroy()
        
        if not coberturas_json:
            return
        
        try:
            coberturas = json.loads(coberturas_json)
            
            for cob in coberturas:
                # Crear fila
                fila = ctk.CTkFrame(self.coberturas_lista_frame, fg_color="transparent")
                fila.pack(fill="x", pady=2)
                
                # Campo de cobertura
                cobertura_entry = ctk.CTkEntry(fila, width=200, height=35, font=ctk.CTkFont(size=11,))
                cobertura_entry.insert(0, cob.get("cobertura", ""))
                cobertura_entry.pack(side="left", padx=5)
                
                # Campo de suma asegurada
                suma_entry = ctk.CTkEntry(fila, width=150, height=35, font=ctk.CTkFont(size=11,))
                suma_entry.insert(0, cob.get("suma_asegurada", ""))
                suma_entry.pack(side="left", padx=5)
                
                # Campo de deducible
                deducible_entry = ctk.CTkEntry(fila, width=120, height=35, font=ctk.CTkFont(size=11,))
                deducible_entry.insert(0, cob.get("deducible", ""))
                deducible_entry.pack(side="left", padx=5)
                
                # Campo de prima
                prima_entry = ctk.CTkEntry(fila, width=100, height=35, font=ctk.CTkFont(size=11,))
                prima_entry.insert(0, cob.get("prima", ""))
                prima_entry.pack(side="left", padx=5)
                
                # Botón para eliminar fila
                btn_eliminar = ctk.CTkButton(
                    fila,
                    text="🗑️",
                    width=40,
                    height=35,
                    fg_color=COLORES["peligro"],
                    hover_color=COLORES["peligro_hover"],
                    command=lambda f=fila: f.destroy()
                )
                btn_eliminar.pack(side="left", padx=5)
                
        except json.JSONDecodeError:
            pass  # Si hay error en JSON, simplemente no cargar nada
    
    def buscar_en_tiempo_real(self, *args):
        """Busca pólizas en tiempo real según el término de búsqueda."""
        termino = self.search_var.get()
        self.cargar_polizas(termino)
    
    def cargar_polizas(self, termino=""):
        """Carga y muestra las pólizas en el panel izquierdo."""
        # Limpiar lista actual
        for widget in self.lista_frame.winfo_children():
            widget.destroy()
        
        # Obtener pólizas
        polizas = self.db.buscar_polizas(termino)
        
        # Actualizar contador
        total = len(polizas)
        self.contador_label.configure(text=f"📊 {total} póliza{'s' if total != 1 else ''} encontrada{'s' if total != 1 else ''}")
        
        # Actualizar header (total de pólizas)
        total_todas = len(self.db.buscar_polizas(""))
        self.total_polizas_label.configure(text=f"📁 {total_todas} Pólizas Registradas")
        
        if not polizas:
            label = ctk.CTkLabel(
                self.lista_frame,
                text="❌ No se encontraron pólizas",
                font=ctk.CTkFont(size=13),
                text_color=COLORES["texto_secundario"]
            )
            label.pack(pady=30)
            return
        
        # Mostrar cada póliza
        for poliza in polizas:
            self.crear_item_poliza(poliza)
    
    def crear_item_poliza(self, poliza):
        """Crea un item de póliza en la lista con diseño mejorado."""
        # Verificar si tiene vencimientos próximos
        tiene_vencimiento = self.db.verificar_vencimientos(poliza)
        
        # Color de fondo según vencimiento
        if tiene_vencimiento:
            fg_color = COLORES["peligro"]
            hover_color = COLORES["peligro_hover"]
        else:
            fg_color = "#334155"
            hover_color = "#475569"
        
        # Preparar texto del item
        icono_advertencia = "⚠️ " if tiene_vencimiento else ""
        nombre = poliza['nombre_completo']
        numero_pol = poliza['numero_poliza'] or 'Sin póliza'
        placas = poliza['placas'] or 'Sin placas'
        marca = poliza['marca'] or ''
        
        # Texto completo
        texto_principal = f"{icono_advertencia}{nombre}"
        texto_info = f"📋 {numero_pol} • 🚗 {placas}"
        if marca:
            texto_info += f" • {marca}"
        
        texto_completo = f"{texto_principal}\n{texto_info}"
        
        # Crear botón clickeable
        btn = ctk.CTkButton(
            self.lista_frame,
            text=texto_completo,
            command=lambda p=poliza: self.seleccionar_poliza(p,),
            fg_color=fg_color,
            hover_color=hover_color,
            text_color="white",
            anchor="w",
            height=75,
            font=ctk.CTkFont(size=12),

            border_width=2,
            border_color=COLORES["borde"]
        )
        btn.pack(fill="x", pady=6, padx=8)
    
    def seleccionar_poliza(self, poliza):
        """Muestra los detalles de una póliza seleccionada."""
        self.poliza_seleccionada_id = poliza['id']
        self.modo_edicion = True
        
        # Llenar formulario
        self.nombre_entry.delete(0, 'end')
        self.nombre_entry.insert(0, poliza.get('nombre_completo', ''))
        self.tipo_cliente_combo.set(poliza.get('tipo_cliente', 'Cliente Normal'))
        
        self.calle_entry.delete(0, 'end')
        self.calle_entry.insert(0, poliza.get('calle', ''))
        
        self.numero_entry.delete(0, 'end')
        self.numero_entry.insert(0, poliza.get('numero', ''))
        
        self.cp_entry.delete(0, 'end')
        self.cp_entry.insert(0, poliza.get('codigo_postal', ''))
        
        colonia = poliza.get('colonia', '')
        self.colonia_combo.set(colonia)
        if colonia and colonia not in self.colonia_combo.cget("values"):
            valores_actuales = list(self.colonia_combo.cget("values"))
            valores_actuales.append(colonia)
            self.colonia_combo.configure(values=valores_actuales)
        
        # Estado y Municipio con ComboBox
        estado = poliza.get('estado', '')
        self.estado_combo.set(estado if estado else '')
        
        # Actualizar municipios y seleccionar
        if estado:
            self.actualizar_municipios(estado)
        
        municipio = poliza.get('municipio', '')
        self.municipio_combo.set(municipio if municipio else '')
        
        self.rfc_entry.delete(0, 'end')
        self.rfc_entry.insert(0, poliza.get('rfc', ''))
        
        self.email_entry.delete(0, 'end')
        self.email_entry.insert(0, poliza.get('email', ''))
        
        self.telefono_entry.delete(0, 'end')
        self.telefono_entry.insert(0, poliza.get('telefono', '') or '')
        
        self.compania_entry.delete(0, 'end')
        self.compania_entry.insert(0, poliza.get('compania_aseguradora', ''))
        
        self.numero_poliza_entry.delete(0, 'end')
        self.numero_poliza_entry.insert(0, poliza.get('numero_poliza', ''))
        
        self.endoso_entry.delete(0, 'end')
        self.endoso_entry.insert(0, poliza.get('endoso', ''))
        
        self.inciso_entry.delete(0, 'end')
        self.inciso_entry.insert(0, poliza.get('inciso', ''))
        
        self.vigencia_inicio_entry.delete(0, 'end')
        self.vigencia_inicio_entry.insert(0, poliza.get('vigencia_inicio', ''))
        
        self.vigencia_fin_entry.delete(0, 'end')
        self.vigencia_fin_entry.insert(0, poliza.get('vigencia_fin', ''))
        
        forma_pago = poliza.get('forma_pago', 'Mensual')
        self.forma_pago_combo.set(forma_pago if forma_pago else 'Mensual')
        
        # Tipo de moneda
        tipo_moneda = poliza.get('tipo_moneda', 'MXN')
        self.tipo_moneda_combo.set(tipo_moneda if tipo_moneda else 'MXN')
        
        self.fecha_vencimiento_entry.delete(0, 'end')
        self.fecha_vencimiento_entry.insert(0, poliza.get('fecha_vencimiento_pago', ''))
        
        # Calcular días restantes
        self.calcular_dias_vencimiento()
        
        self.marca_entry.delete(0, 'end')
        self.marca_entry.insert(0, poliza.get('marca', ''))
        
        self.modelo_entry.delete(0, 'end')
        self.modelo_entry.insert(0, poliza.get('modelo', ''))
        
        self.tipo_version_entry.delete(0, 'end')
        self.tipo_version_entry.insert(0, poliza.get('tipo_version', ''))
        
        self.anio_entry.delete(0, 'end')
        self.anio_entry.insert(0, poliza.get('anio', ''))
        
        self.serie_vin_entry.delete(0, 'end')
        self.serie_vin_entry.insert(0, poliza.get('serie_vin', ''))
        
        self.motor_entry.delete(0, 'end')
        self.motor_entry.insert(0, poliza.get('motor', ''))
        
        self.placas_entry.delete(0, 'end')
        self.placas_entry.insert(0, poliza.get('placas', ''))
        
        # Uso y Servicio del vehículo
        uso = poliza.get('uso', '')
        self.uso_combo.set(uso if uso else '')
        
        servicio = poliza.get('servicio', '')
        self.servicio_combo.set(servicio if servicio else '')
        
        self.ultimo_movimiento_entry.delete(0, 'end')
        self.ultimo_movimiento_entry.insert(0, poliza.get('ultimo_movimiento', ''))
        
        self.prima_entry.delete(0, 'end')
        self.prima_entry.insert(0, str(poliza.get('prima_total', '')))
        
        # Cargar coberturas
        coberturas_json = poliza.get('coberturas_json', '')
        self.cargar_coberturas(coberturas_json)
        
        # Cargar comentarios
        self.comentarios_text.delete("1.0", "end")
        comentarios = poliza.get('comentarios', '')
        if comentarios:
            self.comentarios_text.insert("1.0", comentarios)
        
        # Actualizar título y subtítulo
        self.form_titulo.configure(text=f"✏️ {poliza['nombre_completo']}")
        self.form_subtitulo.configure(text=f"Editando póliza: {poliza['numero_poliza'] or 'Sin número'}")
    
    def nueva_poliza(self):
        """Prepara el formulario para crear una nueva póliza."""
        self.poliza_seleccionada_id = None
        self.modo_edicion = False
        self.limpiar_formulario()
        self.form_titulo.configure(text="🆕 Nueva Póliza")
        self.form_subtitulo.configure(text="Completa la información para registrar una nueva póliza")
    
    def limpiar_formulario(self):
        """Limpia todos los campos del formulario."""
        self.nombre_entry.delete(0, 'end')
        self.tipo_cliente_combo.set("Cliente Normal")
        self.calle_entry.delete(0, 'end')
        self.numero_entry.delete(0, 'end')
        self.cp_entry.delete(0, 'end')
        self.colonia_combo.set("")
        self.colonia_combo.configure(values=[])
        self.municipio_combo.set('')
        self.estado_combo.set('')
        self.rfc_entry.delete(0, 'end')
        self.email_entry.delete(0, 'end')
        self.telefono_entry.delete(0, 'end')
        self.compania_entry.delete(0, 'end')
        self.numero_poliza_entry.delete(0, 'end')
        self.endoso_entry.delete(0, 'end')
        self.inciso_entry.delete(0, 'end')
        self.vigencia_inicio_entry.delete(0, 'end')
        self.vigencia_fin_entry.delete(0, 'end')
        self.forma_pago_combo.set('Mensual')
        self.tipo_moneda_combo.set('MXN')
        self.fecha_vencimiento_entry.delete(0, 'end')
        self.dias_restantes_label.configure(text="")
        self.marca_entry.delete(0, 'end')
        self.modelo_entry.delete(0, 'end')
        self.tipo_version_entry.delete(0, 'end')
        self.anio_entry.delete(0, 'end')
        self.serie_vin_entry.delete(0, 'end')
        self.motor_entry.delete(0, 'end')
        self.placas_entry.delete(0, 'end')
        self.uso_combo.set('')
        self.servicio_combo.set('')
        self.ultimo_movimiento_entry.delete(0, 'end')
        self.prima_entry.delete(0, 'end')
        
        # Limpiar coberturas
        for widget in self.coberturas_lista_frame.winfo_children():
            widget.destroy()
        
        # Limpiar comentarios
        self.comentarios_text.delete("1.0", "end")
    
    def validar_rfc(self, rfc):
        """Valida que el RFC tenga 10 caracteres."""
        return len(rfc) == 10 if rfc else True
    
    def guardar_poliza(self):
        """Guarda o actualiza una póliza."""
        # Validar campos requeridos
        if not self.nombre_entry.get():
            messagebox.showerror("Error", "El nombre completo es obligatorio")
            return
        
        # Validar RFC
        rfc = self.rfc_entry.get()
        if rfc and not self.validar_rfc(rfc):
            messagebox.showerror("Error", "El RFC debe tener exactamente 10 caracteres")
            return
        
        # Validar prima total
        try:
            prima = float(self.prima_entry.get()) if self.prima_entry.get() else 0.0
        except ValueError:
            messagebox.showerror("Error", "La prima total debe ser un número válido")
            return
        
        # CONFIRMACIÓN ANTES DE GUARDAR
        if self.modo_edicion and self.poliza_seleccionada_id:
            # Confirmación para ACTUALIZAR
            respuesta = messagebox.askyesno(
                "Confirmar Actualización",
                f"¿Está seguro de actualizar la póliza de:\n{self.nombre_entry.get()}?\n\nEsta acción modificará los datos existentes.",
                icon='question'
            )
            if not respuesta:
                return
        else:
            # Confirmación para ALTA (nueva póliza)
            respuesta = messagebox.askyesno(
                "Confirmar Alta",
                f"¿Está seguro de crear una nueva póliza para:\n{self.nombre_entry.get()}?",
                icon='question'
            )
            if not respuesta:
                return
        
        # Recopilar datos
        datos = {
            'nombre_completo': self.nombre_entry.get(),
            'calle': self.calle_entry.get(),
            'numero': self.numero_entry.get(),
            'codigo_postal': self.cp_entry.get(),
            'colonia': self.colonia_combo.get(),
            'municipio': self.municipio_combo.get(),
            'estado': self.estado_combo.get(),
            'rfc': rfc,
            'email': self.email_entry.get(),
            'telefono': self.telefono_entry.get(),
            'compania_aseguradora': self.compania_entry.get(),
            'numero_poliza': self.numero_poliza_entry.get(),
            'endoso': self.endoso_entry.get(),
            'inciso': self.inciso_entry.get(),
            'vigencia_inicio': self.vigencia_inicio_entry.get(),
            'vigencia_fin': self.vigencia_fin_entry.get(),
            'forma_pago': self.forma_pago_combo.get(),
            'fecha_vencimiento_pago': self.fecha_vencimiento_entry.get(),
            'tipo_moneda': self.tipo_moneda_combo.get(),
            'marca': self.marca_entry.get(),
            'modelo': self.modelo_entry.get(),
            'tipo_version': self.tipo_version_entry.get(),
            'anio': self.anio_entry.get(),
            'serie_vin': self.serie_vin_entry.get(),
            'motor': self.motor_entry.get(),
            'placas': self.placas_entry.get(),
            'uso': self.uso_combo.get(),
            'servicio': self.servicio_combo.get(),
            'ultimo_movimiento': self.ultimo_movimiento_entry.get(),
            'prima_total': prima,
            'coberturas_json': self.obtener_coberturas(),
            'comentarios': self.comentarios_text.get("1.0", "end-1c")
        }
        
        try:
            if self.modo_edicion and self.poliza_seleccionada_id:
                # Actualizar póliza existente
                self.db.actualizar_poliza(self.poliza_seleccionada_id, datos)
                messagebox.showinfo("Éxito", "Póliza actualizada correctamente")
            else:
                # Crear nueva póliza
                self.db.agregar_poliza(datos)
                messagebox.showinfo("Éxito", "Póliza creada correctamente")
            
            # Recargar lista y limpiar formulario
            self.cargar_polizas()
            self.nueva_poliza()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar: {str(e)}")
    
    def cancelar_edicion(self):
        """Cancela la edición actual."""
        self.nueva_poliza()
    
    def eliminar_poliza(self):
        """Elimina la póliza seleccionada."""
        if not self.poliza_seleccionada_id:
            messagebox.showwarning("Advertencia", "Selecciona una póliza para eliminar")
            return
        
        # Confirmar eliminación
        respuesta = messagebox.askyesno(
            "Confirmar eliminación",
            "¿Estás seguro de que deseas eliminar esta póliza?\nEsta acción no se puede deshacer."
        )
        
        if respuesta:
            try:
                self.db.eliminar_poliza(self.poliza_seleccionada_id)
                messagebox.showinfo("Éxito", "Póliza eliminada correctamente")
                self.cargar_polizas()
                self.nueva_poliza()
            except Exception as e:
                messagebox.showerror("Error", f"Error al eliminar: {str(e)}")
    
    def exportar_excel(self):
        """Exporta la póliza seleccionada a Excel."""
        if not self.poliza_seleccionada_id:
            messagebox.showwarning("Advertencia", "Selecciona una póliza para exportar")
            return
        
        # Obtener datos de la póliza
        poliza = self.db.obtener_poliza(self.poliza_seleccionada_id)
        
        if not poliza:
            messagebox.showerror("Error", "No se pudo obtener la póliza")
            return
        
        # Solicitar ubicación de guardado
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Poliza_{poliza['numero_poliza'] or 'sin_numero'}.xlsx"
        )
        
        if not archivo:
            return
        
        try:
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Póliza"
            
            # Título
            ws['A1'] = "PÓLIZA DE SEGURO DE VEHÍCULO"
            ws['A1'].font = openpyxl.styles.Font(size=16, bold=True)
            
            # Información Personal
            row = 3
            ws[f'A{row}'] = "INFORMACIÓN PERSONAL"
            ws[f'A{row}'].font = openpyxl.styles.Font(size=14, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Nombre Completo:"
            ws[f'B{row}'] = poliza.get('nombre_completo', '')
            row += 1
            
            ws[f'A{row}'] = "Domicilio:"
            ws[f'B{row}'] = f"{poliza.get('calle', '')} {poliza.get('numero', '')}, CP {poliza.get('codigo_postal', '')}"
            row += 1
            
            ws[f'A{row}'] = ""
            ws[f'B{row}'] = f"{poliza.get('colonia', '')}, {poliza.get('municipio', '')}, {poliza.get('estado', '')}"
            row += 1
            
            ws[f'A{row}'] = "RFC:"
            ws[f'B{row}'] = poliza.get('rfc', '')
            row += 1
            
            ws[f'A{row}'] = "Email:"
            ws[f'B{row}'] = poliza.get('email', '')
            row += 2
            
            # Datos de Póliza
            ws[f'A{row}'] = "DATOS DE PÓLIZA"
            ws[f'A{row}'].font = openpyxl.styles.Font(size=14, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Compañía Aseguradora:"
            ws[f'B{row}'] = poliza.get('compania_aseguradora', '')
            row += 1
            
            ws[f'A{row}'] = "Número de Póliza:"
            ws[f'B{row}'] = poliza.get('numero_poliza', '')
            row += 1
            
            ws[f'A{row}'] = "Endoso:"
            ws[f'B{row}'] = poliza.get('endoso', '')
            row += 1
            
            ws[f'A{row}'] = "Inciso:"
            ws[f'B{row}'] = poliza.get('inciso', '')
            row += 1
            
            ws[f'A{row}'] = "Vigencia:"
            ws[f'B{row}'] = f"{poliza.get('vigencia_inicio', '')} al {poliza.get('vigencia_fin', '')}"
            row += 1
            
            ws[f'A{row}'] = "Forma de Pago:"
            ws[f'B{row}'] = poliza.get('forma_pago', '')
            row += 1
            
            ws[f'A{row}'] = "Fecha Vencimiento Pago:"
            ws[f'B{row}'] = poliza.get('fecha_vencimiento_pago', '')
            row += 2
            
            # Datos del Vehículo
            ws[f'A{row}'] = "DATOS DEL VEHÍCULO"
            ws[f'A{row}'].font = openpyxl.styles.Font(size=14, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Marca:"
            ws[f'B{row}'] = poliza.get('marca', '')
            row += 1
            
            ws[f'A{row}'] = "Modelo:"
            ws[f'B{row}'] = poliza.get('modelo', '')
            row += 1
            
            ws[f'A{row}'] = "Tipo/Versión:"
            ws[f'B{row}'] = poliza.get('tipo_version', '')
            row += 1
            
            ws[f'A{row}'] = "Año:"
            ws[f'B{row}'] = poliza.get('anio', '')
            row += 1
            
            ws[f'A{row}'] = "Serie/VIN:"
            ws[f'B{row}'] = poliza.get('serie_vin', '')
            row += 1
            
            ws[f'A{row}'] = "Motor:"
            ws[f'B{row}'] = poliza.get('motor', '')
            row += 1
            
            ws[f'A{row}'] = "Placas:"
            ws[f'B{row}'] = poliza.get('placas', '')
            row += 1
            
            ws[f'A{row}'] = "Prima Total:"
            ws[f'B{row}'] = f"${poliza.get('prima_total', 0):,.2f}"
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40
            
            # Guardar archivo
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Póliza exportada exitosamente a:\n{archivo}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def abrir_cambiar_password(self):
        """Abre ventana para cambiar contraseña del usuario actual."""
        cambio_window = CambiarPasswordWindow(
            self,
            self.db,
            self.usuario_autenticado,
            obligatorio=False
        )
    
    def abrir_gestion_usuarios(self):
        """Abre ventana de gestión de usuarios (solo admin)."""
        if self.usuario_autenticado.get('usuario') == 'admin':
            gestion_window = GestionUsuariosWindow(self, self.db)
        else:
            messagebox.showwarning(
                "Acceso Denegado",
                "Solo el administrador puede gestionar usuarios"
            )
    def abrir_agenda(self):
        """Abre la ventana de Agenda Telefónica."""
        AgendaTelefonicaWindow(self, self.db)
        


    def salir_programa(self):
        """Pide confirmación para salir y cierra la aplicación."""
        respuesta = messagebox.askyesno(
            "Cerrar Sesión",
            "¿Estás seguro de que deseas cerrar sesión y salir del programa?",
            icon='warning'
        )
        if respuesta:
            self.quit()
            self.destroy()
            
    def validar_cp_y_buscar(self, *args):
        """Dispara búsqueda asíncrona cuando el CP tiene 5 dígitos."""
        cp = self.cp_var.get()
        if len(cp) == 5 and cp.isdigit():
            # Iniciar hilo para no bloquear la UI principal
            threading.Thread(target=self.buscar_datos_por_cp, args=(cp,), daemon=True).start()
            
    def buscar_datos_por_cp(self, cp):
        """Thread asíncrono que consulta la API de códigos postales y actualiza form."""
        try:
            url = f"https://api.zippopotam.us/mx/{cp}"
            response = requests.get(url, timeout=5)
            
            if response.status_code == 200:
                data = response.json()
                lugares = data.get("places", [])
                
                if lugares:
                    # Rellenar estado basándonos en la API
                    estado_api = lugares[0].get("state", "").upper()
                    
                    # Rellenar colonias
                    colonias = [lugar.get("place name", "") for lugar in lugares]
                    
                    # Actualizar combo de colonia en el hilo principal
                    self.after(0, lambda: self._actualizar_ui_cp(colonias, estado_api))
                    
        except requests.exceptions.RequestException:
            pass  # Error silencioso en caso de no haber conexión o timeout
            
    def _actualizar_ui_cp(self, colonias, estado_api):
        """Actualiza la UI desde el hilo principal con los resultados del CP."""
        if colonias:
            self.colonia_combo.configure(values=colonias)
            self.colonia_combo.set(colonias[0])
            
        # Intentar hacer match del estado devuelto con nuestro catálogo
        for est in ESTADOS_MEXICO:
            if est.upper() in estado_api or estado_api in est.upper() or self._normalizar_estado(est) == self._normalizar_estado(estado_api):
                self.estado_combo.set(est)
                self.actualizar_municipios(est)
                break
                
    def _normalizar_estado(self, texto):
        """Normaliza texto quitando acentos."""
        import unicodedata
        texto = texto.upper()
        return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def load_config():
    """Carga o crea un archivo de configuración para la red local."""
    import os
    config_path = "config.json"
    default_config = {"db_path": "seguros.db"}
    
    if not os.path.exists(config_path):
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(default_config, f, indent=4)
        except Exception:
            pass
        return default_config
        
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_config

if __name__ == "__main__":
    # Cargar configuración para LAN
    config = load_config()
    db_path = config.get("db_path", "seguros.db")
    
    # Crear ventana temporal para login
    root = ctk.CTk()
    root.withdraw()  # Ocultar ventana principal
    
    # Crear base de datos con la ruta del config
    db = DatabaseManager(db_name=db_path)
    
    # Mostrar ventana de login
    login = LoginWindow(root, db)
    root.wait_window(login)  # Esperar a que se cierre el login
    
    # Verificar si se autenticó
    if login.usuario_autenticado:
        root.destroy()  # Destruir ventana temporal
        
        # Crear aplicación principal con usuario autenticado
        app = SeguroApp(usuario_autenticado=login.usuario_autenticado, db_path=db_path)
        app.mainloop()
    else:
        root.destroy()  # Cerrar si no se autenticó
